#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Steel Wire Receipt Entry Script
Fill customer receipt date and received amount into the customer workbook.

Usage:
  python3 receipt_entry.py -c <customer> -n <sales_contract> -a <amount> [--receipt-date <YYYY.MM.DD>]
"""

import argparse
import os
import sys
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import load_workbook

from workbook_query_utils import (
    SALES_FILE,
    find_first_receipt_pending_row,
    get_contract_receivable_snapshot,
    resolve_sales_sheet_name,
    summarize_receivable_contracts,
)


DATE_FMT = '%Y.%m.%d'


def format_amount(amount):
    return f'{amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP):f}'


def enter_receipt(customer, contract_no, amount, receipt_date=None):
    if not receipt_date:
        receipt_date = datetime.now().strftime(DATE_FMT)

    if amount <= Decimal('0'):
        print('\n❌ Error: Received amount must be greater than 0!')
        return False

    if not os.path.exists(SALES_FILE):
        print(f'\n❌ Error: Sales file not found: {SALES_FILE}')
        return False

    wb_formulas = load_workbook(SALES_FILE, data_only=False)
    wb_values = load_workbook(SALES_FILE, data_only=True)

    try:
        resolved_customer, match_mode = resolve_sales_sheet_name(wb_formulas, customer)
    except KeyError:
        print(f'\n❌ Error: Customer worksheet "{customer}" not found in sales file!')
        print(f'Available worksheets: {", ".join(wb_formulas.sheetnames)}')
        return False

    ws_formulas = wb_formulas[resolved_customer]
    ws_values = wb_values[resolved_customer]

    contracts = summarize_receivable_contracts(
        ws_formulas=ws_formulas,
        ws_values=ws_values,
        settled='all',
        contract_no=contract_no,
    )
    if not contracts:
        print(f'\n❌ Error: Sales contract "{contract_no}" not found for customer "{customer}"!')
        return False

    contract = contracts[0]
    balance = get_contract_receivable_snapshot(ws_formulas, ws_values, contract_no)
    current_unreceived_amount = balance['unreceived_amount'] if balance else None
    if current_unreceived_amount is None:
        print(f'\n❌ Error: Unable to read balance snapshot for contract "{contract_no}"!')
        return False

    if current_unreceived_amount == Decimal('0.00'):
        print(f'\n❌ Error: Sales contract "{contract_no}" is already fully settled!')
        return False

    allowed_amount = abs(current_unreceived_amount)
    if amount > allowed_amount:
        print(f'\n❌ Error: Receipt amount exceeds current outstanding balance!')
        print(f'   Receipt amount: {format_amount(amount)}')
        print(f'   Outstanding balance: {format_amount(allowed_amount)}')
        return False

    target_row = find_first_receipt_pending_row(ws_formulas, str(contract_no))
    if target_row is None:
        print(f'\n❌ Error: Receipt entries for contract "{contract_no}" exceed the available truck rows!')
        print('Please handle this contract manually.')
        return False

    ws_formulas.cell(row=target_row, column=21, value=receipt_date)
    ws_formulas.cell(row=target_row, column=22, value=float(amount))
    wb_formulas.save(SALES_FILE)

    print('\n=== Receipt Entry ===')
    print(f'Customer: {customer}')
    if resolved_customer != customer:
        print(f'Resolved Customer Sheet: {resolved_customer} ({match_mode})')
    print(f'Sales Contract: {contract_no}')
    print(f'Target Row: {target_row}')
    print(f'Balance Row: {balance["row"]}')
    print(f'Receipt Date: {receipt_date}')
    print(f'Received Amount: {format_amount(amount)}')
    print(f'Contract Sales Total: {format_amount(balance["total_sales_amount"])}')
    print(f'Contract Received Total Before Write: {format_amount(balance["received_amount"])}')
    print(f'Outstanding Balance Before Write: {format_amount(allowed_amount)}')
    print(f'\n✅ Sales file updated: {SALES_FILE}')
    print('✅ Receipt entry completed!')
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Steel Wire Receipt Entry Script',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python3 receipt_entry.py -c "东莞建安" -n 2026032401 -a 50000
  python3 receipt_entry.py -c "东莞建安" -n 2026032401 -a 50000 --receipt-date 2026.03.24
        '''
    )

    parser.add_argument('-c', '--customer', required=True, help='Customer worksheet name in the sales workbook.')
    parser.add_argument('-n', '--contract', required=True, help='Sales contract number.')
    parser.add_argument('-a', '--amount', required=True, type=Decimal, help='Received amount for this receipt entry.')
    parser.add_argument('--receipt-date', help='Receipt date, format YYYY.MM.DD. Defaults to today.')

    arguments = parser.parse_args()
    success = enter_receipt(
        customer=arguments.customer,
        contract_no=arguments.contract,
        amount=arguments.amount,
        receipt_date=arguments.receipt_date,
    )
    sys.exit(0 if success else 1)
