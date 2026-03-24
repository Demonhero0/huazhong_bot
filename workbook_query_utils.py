#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import os
import re
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime


ORDER_FILE = os.path.join(os.path.dirname(__file__), '线材供应商提货明细龙虾版.xlsx')
SALES_FILE = os.path.join(os.path.dirname(__file__), '线材客户送货明细龙虾版.xlsx')


def is_blank(value):
    return value is None or value == ''


def to_decimal(value, default='0'):
    if value is None or value == '':
        return Decimal(default)
    text = str(value).strip()
    if text.startswith('='):
        return Decimal(default)
    return Decimal(text).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def parse_date_text(value):
    if is_blank(value):
        return None
    text = str(value).strip()
    return datetime.strptime(text, '%Y.%m.%d')


def load_order_workbook():
    return load_workbook(ORDER_FILE)


def load_sales_workbook():
    return load_workbook(SALES_FILE)


def get_order_sheet(wb, supplier):
    if supplier not in wb.sheetnames:
        raise KeyError(f'Supplier worksheet not found: {supplier}')
    return wb[supplier]


def get_sales_sheet(wb, customer):
    resolved_name, _ = resolve_sales_sheet_name(wb, customer)
    return wb[resolved_name]


def normalize_customer_name(name):
    text = str(name or '').strip()
    text = re.sub(r'[\s()（）\-.·,，]', '', text)
    for token in ['有限责任公司', '股份有限公司', '有限公司', '责任公司', '公司']:
        if text.endswith(token):
            text = text[:-len(token)]
    text = text.replace('省', '').replace('市', '')
    return text


def resolve_sales_sheet_name(wb, customer):
    """
    Resolve customer input to an existing sales worksheet name.
    Matching order:
    1. exact sheet name
    2. normalized unique match
    3. normalized containment unique match
    """
    if customer in wb.sheetnames:
        return customer, 'exact'

    normalized_input = normalize_customer_name(customer)
    if not normalized_input:
        raise KeyError('Customer worksheet not found: empty customer name')

    normalized_matches = []
    containment_matches = []
    for sheet_name in wb.sheetnames:
        normalized_sheet = normalize_customer_name(sheet_name)
        if normalized_sheet == normalized_input:
            normalized_matches.append(sheet_name)
        elif normalized_input in normalized_sheet or normalized_sheet in normalized_input:
            containment_matches.append(sheet_name)

    if len(normalized_matches) == 1:
        return normalized_matches[0], 'normalized'
    if len(normalized_matches) > 1:
        raise KeyError(
            f'Customer worksheet name "{customer}" is ambiguous. Candidates: {", ".join(normalized_matches)}'
        )

    if len(containment_matches) == 1:
        return containment_matches[0], 'containment'
    if len(containment_matches) > 1:
        raise KeyError(
            f'Customer worksheet name "{customer}" is ambiguous. Candidates: {", ".join(containment_matches)}'
        )

    raise KeyError(f'Customer worksheet not found: {customer}')


def iter_contract_rows(ws, contract_no):
    for row in range(5, ws.max_row + 1):
        if str(ws.cell(row=row, column=2).value) == str(contract_no):
            yield row


def get_contract_rows(ws, contract_no):
    return list(iter_contract_rows(ws, contract_no))


def get_contract_balance_snapshot(ws_values, contract_no):
    rows = get_contract_rows(ws_values, contract_no)
    if not rows:
        return None
    last_row = rows[-1]
    return {
        'row': last_row,
        'received_amount': to_decimal(ws_values.cell(row=last_row, column=22).value),
        'unreceived_amount': to_decimal(ws_values.cell(row=last_row, column=23).value),
    }


def resolve_sales_amount(ws_formulas, ws_values, row):
    cached_amount = ws_values.cell(row=row, column=20).value
    if not is_blank(cached_amount):
        return to_decimal(cached_amount)
    received_weight = to_decimal(ws_formulas.cell(row=row, column=18).value)
    unit_price = to_decimal(ws_formulas.cell(row=row, column=19).value)
    return (received_weight * unit_price).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def get_contract_receivable_snapshot(ws_formulas, ws_values, contract_no):
    rows = get_contract_rows(ws_formulas, contract_no)
    if not rows:
        return None

    total_sales_amount = Decimal('0.00')
    total_received_amount = Decimal('0.00')
    for row in rows:
        total_sales_amount += resolve_sales_amount(ws_formulas, ws_values, row)
        total_received_amount += to_decimal(ws_formulas.cell(row=row, column=22).value)

    last_row = rows[-1]
    outstanding_amount = (total_received_amount - total_sales_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return {
        'row': last_row,
        'rows': rows,
        'total_sales_amount': total_sales_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'received_amount': total_received_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'unreceived_amount': outstanding_amount,
    }


def summarize_order_contracts(ws):
    contracts = {}
    for row in range(5, ws.max_row + 1):
        contract_no = ws.cell(row=row, column=2).value
        if not contract_no:
            continue

        contract = contracts.setdefault(contract_no, {
            'contract_no': contract_no,
            'start_row': row,
            'end_row': row,
            'total_rows': 0,
            'transport': ws.cell(row=row, column=7).value,
            'brand': ws.cell(row=row, column=4).value,
            'spec': ws.cell(row=row, column=6).value,
            'unit_price': ws.cell(row=row, column=11).value,
            'order_date': ws.cell(row=row, column=3).value,
            'pending_pickup_rows': [],
            'completed_pickup_rows': [],
            'pending_customer_rows': [],
            'assigned_customer_rows': [],
        })
        contract['end_row'] = row
        contract['total_rows'] += 1

        pickup_date = ws.cell(row=row, column=8).value
        truck_no = ws.cell(row=row, column=9).value
        factory_weight = ws.cell(row=row, column=10).value
        customer = ws.cell(row=row, column=16).value

        if is_blank(pickup_date) and is_blank(truck_no) and is_blank(factory_weight):
            contract['pending_pickup_rows'].append(row)
        else:
            contract['completed_pickup_rows'].append(row)

        if is_blank(customer):
            contract['pending_customer_rows'].append(row)
        else:
            contract['assigned_customer_rows'].append(row)

    result = []
    for contract in contracts.values():
        result.append({
            'contract_no': contract['contract_no'],
            'start_row': contract['start_row'],
            'end_row': contract['end_row'],
            'total_rows': contract['total_rows'],
            'transport': contract['transport'],
            'brand': contract['brand'],
            'spec': contract['spec'],
            'unit_price': contract['unit_price'],
            'order_date': contract['order_date'],
            'pending_pickup_rows': contract['pending_pickup_rows'],
            'completed_pickup_rows': contract['completed_pickup_rows'],
            'pending_customer_rows': contract['pending_customer_rows'],
            'assigned_customer_rows': contract['assigned_customer_rows'],
            'pending_pickup_count': len(contract['pending_pickup_rows']),
            'completed_pickup_count': len(contract['completed_pickup_rows']),
            'pending_customer_count': len(contract['pending_customer_rows']),
            'assigned_customer_count': len(contract['assigned_customer_rows']),
        })
    return sorted(result, key=lambda item: str(item['contract_no']))


def summarize_sales_contracts(ws):
    contracts = {}
    for row in range(5, ws.max_row + 1):
        contract_no = ws.cell(row=row, column=2).value
        if not contract_no:
            continue

        contract = contracts.setdefault(contract_no, {
            'contract_no': contract_no,
            'start_row': row,
            'end_row': row,
            'total_rows': 0,
            'supplier': ws.cell(row=row, column=12).value,
            'brand': ws.cell(row=row, column=4).value,
            'spec': ws.cell(row=row, column=5).value,
            'sales_date': ws.cell(row=row, column=3).value,
            'sell_price': ws.cell(row=row, column=19).value,
            'pending_delivery_rows': [],
            'completed_delivery_rows': [],
        })
        contract['end_row'] = row
        contract['total_rows'] += 1

        delivery_date = ws.cell(row=row, column=8).value
        truck_no = ws.cell(row=row, column=16).value
        factory_weight = ws.cell(row=row, column=17).value
        received_weight = ws.cell(row=row, column=18).value

        if is_blank(delivery_date) and is_blank(truck_no) and is_blank(factory_weight) and is_blank(received_weight):
            contract['pending_delivery_rows'].append(row)
        else:
            contract['completed_delivery_rows'].append(row)

    result = []
    for contract in contracts.values():
        result.append({
            'contract_no': contract['contract_no'],
            'start_row': contract['start_row'],
            'end_row': contract['end_row'],
            'total_rows': contract['total_rows'],
            'supplier': contract['supplier'],
            'brand': contract['brand'],
            'spec': contract['spec'],
            'sales_date': contract['sales_date'],
            'sell_price': contract['sell_price'],
            'pending_delivery_rows': contract['pending_delivery_rows'],
            'completed_delivery_rows': contract['completed_delivery_rows'],
            'pending_delivery_count': len(contract['pending_delivery_rows']),
            'completed_delivery_count': len(contract['completed_delivery_rows']),
        })
    return sorted(result, key=lambda item: str(item['contract_no']))


def find_pending_order_rows(ws, contract_no):
    rows = []
    for row in iter_contract_rows(ws, contract_no):
        pickup_date = ws.cell(row=row, column=8).value
        truck_no = ws.cell(row=row, column=9).value
        factory_weight = ws.cell(row=row, column=10).value
        if not (is_blank(pickup_date) and is_blank(truck_no) and is_blank(factory_weight)):
            continue
        rows.append({
            'row': row,
            'brand': ws.cell(row=row, column=4).value,
            'spec': ws.cell(row=row, column=6).value,
            'transport': ws.cell(row=row, column=7).value,
            'customer': ws.cell(row=row, column=16).value,
            'sell_price': ws.cell(row=row, column=17).value,
            'delivery_mode': ws.cell(row=row, column=18).value,
            'dock': ws.cell(row=row, column=5).value,
            'pickup_date': pickup_date,
            'truck_no': truck_no,
            'factory_weight': factory_weight,
        })
    return rows


def find_pending_sales_rows(ws, contract_no):
    rows = []
    for row in iter_contract_rows(ws, contract_no):
        delivery_date = ws.cell(row=row, column=8).value
        truck_no = ws.cell(row=row, column=16).value
        factory_weight = ws.cell(row=row, column=17).value
        received_weight = ws.cell(row=row, column=18).value
        if not (is_blank(delivery_date) and is_blank(truck_no) and is_blank(factory_weight) and is_blank(received_weight)):
            continue
        rows.append({
            'row': row,
            'supplier': ws.cell(row=row, column=12).value,
            'brand': ws.cell(row=row, column=4).value,
            'spec': ws.cell(row=row, column=5).value,
            'delivery_mode': ws.cell(row=row, column=14).value,
            'sell_price': ws.cell(row=row, column=19).value,
            'delivery_date': delivery_date,
            'truck_no': truck_no,
            'factory_weight': factory_weight,
            'received_weight': received_weight,
        })
    return rows


def summarize_receivable_contracts(ws_formulas, ws_values, date_from=None, date_to=None, settled='no', contract_no=None):
    contracts = {}
    for row in range(5, ws_formulas.max_row + 1):
        current_contract = ws_formulas.cell(row=row, column=2).value
        if is_blank(current_contract):
            continue
        current_contract = str(current_contract)
        if contract_no and current_contract != str(contract_no):
            continue

        sales_date = ws_formulas.cell(row=row, column=3).value
        sales_dt = parse_date_text(sales_date) if not is_blank(sales_date) else None
        if date_from and sales_dt and sales_dt < date_from:
            continue
        if date_to and sales_dt and sales_dt > date_to:
            continue

        contract = contracts.setdefault(current_contract, {
            'contract_no': current_contract,
            'sales_date': sales_date,
            'start_row': row,
            'end_row': row,
            'total_rows': 0,
            'supplier': ws_formulas.cell(row=row, column=12).value,
            'brand': ws_formulas.cell(row=row, column=4).value,
            'spec': ws_formulas.cell(row=row, column=5).value,
            'sell_price': ws_formulas.cell(row=row, column=19).value,
            'receipt_filled_rows': [],
            'receipt_pending_rows': [],
        })
        contract['end_row'] = row
        contract['total_rows'] += 1

        receipt_date = ws_formulas.cell(row=row, column=21).value
        received_amount = ws_formulas.cell(row=row, column=22).value
        if is_blank(receipt_date) and is_blank(received_amount):
            contract['receipt_pending_rows'].append(row)
        else:
            contract['receipt_filled_rows'].append(row)

    result = []
    for contract in contracts.values():
        balance = get_contract_receivable_snapshot(ws_formulas, ws_values, contract['contract_no'])
        current_received_amount = balance['received_amount'] if balance else Decimal('0.00')
        current_unreceived_amount = balance['unreceived_amount'] if balance else Decimal('0.00')
        is_settled = current_unreceived_amount == Decimal('0.00')
        if settled == 'yes' and not is_settled:
            continue
        if settled == 'no' and is_settled:
            continue
        result.append({
            'contract_no': contract['contract_no'],
            'sales_date': contract['sales_date'],
            'start_row': contract['start_row'],
            'end_row': contract['end_row'],
            'total_rows': contract['total_rows'],
            'supplier': contract['supplier'],
            'brand': contract['brand'],
            'spec': contract['spec'],
            'sell_price': contract['sell_price'],
            'receipt_filled_rows': contract['receipt_filled_rows'],
            'receipt_pending_rows': contract['receipt_pending_rows'],
            'receipt_filled_count': len(contract['receipt_filled_rows']),
            'receipt_pending_count': len(contract['receipt_pending_rows']),
            'balance_row': balance['row'] if balance else None,
            'current_received_amount': f'{current_received_amount:.2f}',
            'current_unreceived_amount': f'{current_unreceived_amount:.2f}',
            'is_settled': is_settled,
        })
    return sorted(result, key=lambda item: str(item['contract_no']))


def find_first_receipt_pending_row(ws, contract_no):
    for row in iter_contract_rows(ws, contract_no):
        receipt_date = ws.cell(row=row, column=21).value
        received_amount = ws.cell(row=row, column=22).value
        if is_blank(receipt_date) and is_blank(received_amount):
            return row
    return None
