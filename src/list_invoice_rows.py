#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import sys

from openpyxl import load_workbook

from workbook_query_utils import SALES_FILE, parse_date_text, resolve_sales_sheet_name, summarize_invoice_rows


def main():
    parser = argparse.ArgumentParser(description='List invoice-status rows from the sales workbook as JSON.')
    parser.add_argument('-c', '--customer', help='Customer worksheet name. Optional; omit to scan all customer sheets.')
    parser.add_argument('-n', '--contract', help='Sales contract number. Optional.')
    parser.add_argument('--date-from', help='Delivery date start, format YYYY.MM.DD.')
    parser.add_argument('--date-to', help='Delivery date end, format YYYY.MM.DD.')
    parser.add_argument(
        '--invoiced',
        default='no',
        choices=['yes', 'no', 'all'],
        help='Filter by invoice status. Default: no.',
    )
    args = parser.parse_args()

    try:
        wb_formulas = load_workbook(SALES_FILE, data_only=False)
        wb_values = load_workbook(SALES_FILE, data_only=True)
    except FileNotFoundError as exc:
        print(json.dumps({'ok': False, 'error': str(exc), 'workbook': SALES_FILE}, ensure_ascii=False, indent=2))
        sys.exit(1)

    date_from = parse_date_text(args.date_from)
    date_to = parse_date_text(args.date_to)

    rows = []
    scanned_customers = []
    resolved_customer = None
    match_mode = None

    try:
        if args.customer:
            resolved_customer, match_mode = resolve_sales_sheet_name(wb_formulas, args.customer)
            customer_rows = summarize_invoice_rows(
                ws_formulas=wb_formulas[resolved_customer],
                ws_values=wb_values[resolved_customer],
                customer_name=resolved_customer,
                date_from=date_from,
                date_to=date_to,
                invoiced=args.invoiced,
                contract_no=args.contract,
            )
            rows.extend(customer_rows)
            scanned_customers.append(resolved_customer)
        else:
            for sheet_name in wb_formulas.sheetnames:
                customer_rows = summarize_invoice_rows(
                    ws_formulas=wb_formulas[sheet_name],
                    ws_values=wb_values[sheet_name],
                    customer_name=sheet_name,
                    date_from=date_from,
                    date_to=date_to,
                    invoiced=args.invoiced,
                    contract_no=args.contract,
                )
                if customer_rows:
                    scanned_customers.append(sheet_name)
                    rows.extend(customer_rows)
    except KeyError as exc:
        print(json.dumps({'ok': False, 'error': str(exc), 'workbook': SALES_FILE}, ensure_ascii=False, indent=2))
        sys.exit(1)

    rows.sort(key=lambda item: (str(item['delivery_date']), item['customer'], str(item['contract_no']), item['row']))

    print(json.dumps({
        'ok': True,
        'workbook': SALES_FILE,
        'customer': args.customer,
        'resolved_customer': resolved_customer,
        'match_mode': match_mode,
        'filters': {
            'contract': args.contract,
            'date_from': args.date_from,
            'date_to': args.date_to,
            'invoiced': args.invoiced,
        },
        'scanned_customers': scanned_customers,
        'rows': rows,
        'row_count': len(rows),
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
