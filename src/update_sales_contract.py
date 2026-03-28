#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Update customer sales-contract fields in the sales workbook.

Usage:
  python3 src/update_sales_contract.py -c <customer> -n <sales_contract> --set <field=value> [--set <field=value> ...]
"""

import argparse
import os
import sys

from openpyxl import load_workbook

from workbook_query_utils import SALES_FILE, get_contract_rows, resolve_sales_sheet_name


FIELD_MAP = {
    'sales_date': 3,
    'brand': 4,
    'spec': 5,
    'benchmark': 6,
    'price_diff': 7,
    'delivery_date': 8,
    'fleet': 9,
    'freight': 10,
    'freight_tax': 11,
    'supplier': 12,
    'order_price': 13,
    'transport_mode': 14,
    'delivery_mode': 14,
    'dock': 15,
    'truck_no': 16,
    'factory_weight': 17,
    'received_weight': 18,
    'sell_price': 19,
    'receipt_date': 21,
    'received_amount': 22,
}


def parse_assignment(text):
    if '=' not in text:
        raise ValueError(f'Invalid --set value "{text}". Expected field=value.')
    field, value = text.split('=', 1)
    field = field.strip()
    value = value.strip()
    if field not in FIELD_MAP:
        raise ValueError(f'Unsupported field "{field}". Supported fields: {", ".join(sorted(FIELD_MAP))}')
    return field, value


def convert_value(field, value):
    if value == '':
        return ''
    if field in {'factory_weight', 'received_weight', 'order_price', 'sell_price', 'freight', 'received_amount', 'price_diff'}:
        try:
            return float(value)
        except ValueError as exc:
            raise ValueError(f'Field "{field}" expects a numeric value, got "{value}".') from exc
    return value


def update_sales_contract(customer, contract_no, assignments):
    if not os.path.exists(SALES_FILE):
        print(f'\n❌ Error: Sales file not found: {SALES_FILE}')
        return False

    wb = load_workbook(SALES_FILE)
    try:
        resolved_customer, match_mode = resolve_sales_sheet_name(wb, customer)
    except KeyError:
        print(f'\n❌ Error: Customer worksheet "{customer}" not found in sales file!')
        print(f'Available worksheets: {", ".join(wb.sheetnames)}')
        return False

    ws = wb[resolved_customer]
    rows = get_contract_rows(ws, contract_no)
    if not rows:
        print(f'\n❌ Error: Sales contract "{contract_no}" not found for customer "{customer}"!')
        return False

    parsed = []
    for assignment in assignments:
        field, raw_value = parse_assignment(assignment)
        value = convert_value(field, raw_value)
        parsed.append((field, value))

    for row in rows:
        for field, value in parsed:
            ws.cell(row=row, column=FIELD_MAP[field], value=value)

    wb.save(SALES_FILE)

    print('\n=== Sales Contract Update ===')
    print(f'Customer: {customer}')
    if resolved_customer != customer:
        print(f'Resolved Customer Sheet: {resolved_customer} ({match_mode})')
    print(f'Sales Contract: {contract_no}')
    print(f'Rows Updated: {rows[0]}-{rows[-1]}')
    for field, value in parsed:
        print(f'  {field} = {value}')
    print(f'\n✅ Sales file updated: {SALES_FILE}')
    print('✅ Sales contract update completed!')
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Update customer sales-contract fields in the sales workbook.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python3 src/update_sales_contract.py -c "东莞建安" -n 2026032401 --set benchmark=迁安自提 --set price_diff=70
  python3 src/update_sales_contract.py -c "东莞建安" -n 2026032401 --set delivery_mode=自提 --set sell_price=3420
        '''
    )
    parser.add_argument('-c', '--customer', required=True, help='Customer worksheet name.')
    parser.add_argument('-n', '--contract', required=True, help='Sales contract number.')
    parser.add_argument('--set', action='append', dest='assignments', required=True, help='Field assignment in the form field=value.')
    args = parser.parse_args()

    try:
        success = update_sales_contract(
            customer=args.customer,
            contract_no=args.contract,
            assignments=args.assignments,
        )
    except ValueError as exc:
        print(f'\n❌ Error: {exc}')
        sys.exit(1)

    sys.exit(0 if success else 1)
