#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Issue customer statements from the sales workbook.

Usage:
  python3 statement_issue.py --customer <customer> [--date-from <YYYY.MM.DD>] [--date-to <YYYY.MM.DD>] [--contract <contract_no>] [--paid yes|no|all] [--statement-date <YYYY.MM.DD>] [--multi-contract-mode summary|split]
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

from workbook_query_utils import SALES_FILE, resolve_sales_sheet_name


SELLER_NAME = '东莞市华众供应链管理有限公司'
OUTPUT_ROOT = Path(__file__).resolve().parent.parent / 'statements'
DATE_FMT = '%Y.%m.%d'


def parse_date(date_text: str | None) -> datetime | None:
    if not date_text:
        return None
    return datetime.strptime(str(date_text).strip(), DATE_FMT)


def normalize_date_text(date_text: str | None) -> str | None:
    if not date_text:
        return None
    return parse_date(date_text).strftime(DATE_FMT)


def money(value: Any) -> Decimal:
    if value is None or value == '':
        return Decimal('0')
    text = str(value).strip()
    if text.startswith('='):
        return Decimal('0')
    return Decimal(text).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def format_money(value: Decimal) -> str:
    return f'{value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP):f}'


def format_cn_date(date_text: str | None) -> str:
    normalized = normalize_date_text(date_text)
    if not normalized:
        return ''
    dt = parse_date(normalized)
    return dt.strftime('%Y年%m月%d日')


def value_is_blank(value: Any) -> bool:
    return value is None or value == ''


def contract_is_paid(rows: list[dict[str, Any]]) -> bool:
    for row in rows:
        if not (value_is_blank(row['receipt_date']) and value_is_blank(row['received_amount'])):
            return True
    return False


def amount_to_uppercase(amount: Decimal) -> str:
    digits = '零壹贰叁肆伍陆柒捌玖'
    units_small = ['', '拾', '佰', '仟']
    units_big = ['', '万', '亿', '兆']

    amount = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    integer = int(amount)
    fraction = int((amount - Decimal(integer)) * 100)

    if integer == 0:
        integer_part = '零元'
    else:
        parts = []
        group_index = 0
        zero_pending = False
        while integer > 0:
            group = integer % 10000
            integer //= 10000
            if group == 0:
                if parts:
                    zero_pending = True
                group_index += 1
                continue

            group_text = ''
            inner_zero = False
            for idx in range(4):
                digit = group % 10
                group //= 10
                if digit == 0:
                    if group_text:
                        inner_zero = True
                    continue
                if inner_zero:
                    group_text = '零' + group_text
                    inner_zero = False
                group_text = digits[digit] + units_small[idx] + group_text

            if zero_pending:
                group_text = '零' + group_text
                zero_pending = False
            group_text += units_big[group_index]
            parts.insert(0, group_text)
            group_index += 1
        integer_part = ''.join(parts) + '元'

    if fraction == 0:
        return integer_part + '整'

    jiao = fraction // 10
    fen = fraction % 10
    frac_text = ''
    if jiao:
        frac_text += digits[jiao] + '角'
    if fen:
        if not jiao:
            frac_text += '零'
        frac_text += digits[fen] + '分'
    return integer_part + frac_text


def safe_folder_name(text: str) -> str:
    invalid = '<>:"/\\|?*'
    for ch in invalid:
        text = text.replace(ch, '_')
    return text.strip()


def build_output_folder_label(customer_full_name: str, contracts: list[dict[str, Any]]) -> str:
    sales_dates = [c['sales_date'] for c in contracts if c.get('sales_date')]
    if not sales_dates:
        suffix = datetime.now().strftime(DATE_FMT)
    elif len(set(sales_dates)) == 1:
        suffix = sales_dates[0]
    else:
        suffix = f'{min(sales_dates)}_{max(sales_dates)}'
    return safe_folder_name(f'{customer_full_name}_{suffix}')


def build_output_file_stem(customer_full_name: str, contracts: list[dict[str, Any]]) -> str:
    return f'{build_output_folder_label(customer_full_name, contracts)}_对账单'


def make_output_dir(customer_full_name: str, contracts: list[dict[str, Any]]) -> Path:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    base_name = build_output_folder_label(customer_full_name, contracts)
    output_dir = OUTPUT_ROOT / base_name
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def build_row_record(ws, row: int) -> dict[str, Any]:
    brand = ws.cell(row=row, column=4).value
    spec = ws.cell(row=row, column=5).value
    return {
        'row': row,
        'sales_contract': ws.cell(row=row, column=2).value,
        'sales_date': ws.cell(row=row, column=3).value,
        'brand': brand,
        'spec': spec,
        'truck_no': ws.cell(row=row, column=16).value,
        'received_weight': ws.cell(row=row, column=18).value,
        'unit_price': ws.cell(row=row, column=19).value,
        'amount': money(ws.cell(row=row, column=20).value),
        'receipt_date': ws.cell(row=row, column=21).value,
        'received_amount': ws.cell(row=row, column=22).value,
        'unreceived_amount': ws.cell(row=row, column=23).value,
        'remark': f'{brand or ""} {spec or ""}'.strip(),
    }


def collect_matching_contracts(
    customer_input: str,
    contract: str | None,
    date_from: str | None,
    date_to: str | None,
    paid: str,
) -> tuple[str, str, list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb_names = load_workbook(SALES_FILE, data_only=False)
    resolved_customer, match_mode = resolve_sales_sheet_name(wb_names, customer_input)
    ws = wb_names[resolved_customer]
    wb_values = load_workbook(SALES_FILE, data_only=True)
    ws_values = wb_values[resolved_customer]

    date_from_dt = parse_date(date_from)
    date_to_dt = parse_date(date_to)

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in range(5, ws.max_row + 1):
        contract_no = ws.cell(row=row, column=2).value
        if value_is_blank(contract_no):
            continue
        contract_str = str(contract_no)
        if contract and contract_str != str(contract):
            continue

        record = build_row_record(ws, row)
        amount_value = ws_values.cell(row=row, column=20).value
        if value_is_blank(amount_value):
            received_weight = Decimal(str(record['received_weight'] or 0))
            unit_price = Decimal(str(record['unit_price'] or 0))
            amount_value = (received_weight * unit_price).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        record['amount'] = money(amount_value)
        sales_date_dt = parse_date(record['sales_date'])
        if date_from_dt and sales_date_dt and sales_date_dt < date_from_dt:
            continue
        if date_to_dt and sales_date_dt and sales_date_dt > date_to_dt:
            continue
        grouped.setdefault(contract_str, []).append(record)

    contracts = []
    for contract_no, rows in sorted(grouped.items(), key=lambda item: item[0]):
        paid_state = contract_is_paid(rows)
        if paid == 'yes' and not paid_state:
            continue
        if paid == 'no' and paid_state:
            continue
        rows.sort(key=lambda item: item['row'])
        total_received_weight = sum(Decimal(str(item['received_weight'] or 0)) for item in rows)
        total_amount = sum(item['amount'] for item in rows)
        last_row = rows[-1]
        contracts.append({
            'contract_no': contract_no,
            'sales_date': rows[0]['sales_date'],
            'rows': rows,
            'paid': paid_state,
            'truck_count': len(rows),
            'total_received_weight': total_received_weight.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
            'total_amount': total_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
            'received_amount_total': money(last_row['received_amount']),
            'unreceived_amount_total': money(last_row['unreceived_amount']),
        })

    return resolved_customer, match_mode, contracts


def build_statement_payload(
    customer_full_name: str,
    statement_date: str,
    contracts: list[dict[str, Any]],
    mode: str,
) -> dict[str, Any]:
    all_rows = []
    for contract in contracts:
        all_rows.extend(contract['rows'])

    total_received_weight = sum(Decimal(str(row['received_weight'] or 0)) for row in all_rows)
    total_amount = sum(row['amount'] for row in all_rows)
    received_amount_total = sum(contract['received_amount_total'] for contract in contracts)
    unreceived_amount_total = sum(contract['unreceived_amount_total'] for contract in contracts)

    if mode == 'summary':
        contract_display = '汇总对账'
        signing_date_display = ''
    else:
        contract_display = contracts[0]['contract_no']
        signing_date_display = contracts[0]['sales_date'] or ''

    display_rows = []
    for idx, row in enumerate(all_rows, 1):
        display_rows.append({
            'index': idx,
            'truck_no': row['truck_no'] or '',
            'received_weight': f'{Decimal(str(row["received_weight"] or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP):f}',
            'unit_price': f'{Decimal(str(row["unit_price"] or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP):f}',
            'amount': format_money(row['amount']),
            'remark': row['remark'],
            'sales_date': row['sales_date'] or '',
            'contract_no': row['sales_contract'],
        })

    return {
        'seller_name': SELLER_NAME,
        'buyer_name': customer_full_name,
        'statement_date': statement_date,
        'contract_display': contract_display,
        'contract_signing_date': signing_date_display,
        'contracts': contracts,
        'mode': mode,
        'rows': display_rows,
        'truck_count': len(display_rows),
        'total_received_weight': total_received_weight.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'total_amount': total_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'total_amount_uppercase': amount_to_uppercase(total_amount),
        'received_amount_total': received_amount_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'unreceived_amount_total': unreceived_amount_total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
    }


def render_excel(statement: dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = '对账单'
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    title_font = Font(size=18, bold=True)
    subtitle_font = Font(size=16, bold=True)
    text_font = Font(size=10)
    note_font = Font(size=10, color='FF0000', bold=True)

    outstanding_amount = (statement['total_amount'] - statement['received_amount_total']).quantize(
        Decimal('0.01'),
        rounding=ROUND_HALF_UP,
    )
    opening_outstanding = Decimal('0.00')

    ws.merge_cells('A1:H1')
    ws['A1'] = statement['seller_name']
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells('A2:H2')
    ws['A2'] = '线材对账单'
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = center

    ws.merge_cells('A3:H3')
    ws['A3'] = f'购货单位：{statement["buyer_name"]}'
    ws['A3'].alignment = left
    ws['A3'].font = text_font

    ws.merge_cells('A4:A5')
    ws.merge_cells('B4:B5')
    ws.merge_cells('C4:C5')
    ws.merge_cells('D4:D5')
    ws.merge_cells('E4:E5')
    ws.merge_cells('F4:F5')
    ws.merge_cells('G4:G5')
    ws.merge_cells('H4:H5')
    ws['A4'] = '序号'
    ws['B4'] = '日期'
    ws['C4'] = '名称规格'
    ws['D4'] = '重量\n(吨)'
    ws['E4'] = '单价\n(元/吨)'
    ws['F4'] = '货款金额\n(元)'
    ws['G4'] = '已收货款金额'
    ws['H4'] = '未收货款金额\n(元)'

    for row in range(4, 6):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center
            cell.font = text_font

    ws.merge_cells('A6:C6')
    ws['A6'] = '未付款金额'
    ws['A6'].alignment = center
    ws['H6'] = float(opening_outstanding)
    ws['H6'].number_format = '#,##0.00'
    for col in range(1, 9):
        cell = ws.cell(row=6, column=col)
        cell.border = border
        cell.alignment = center
        cell.font = text_font

    row_no = 7
    running_outstanding = opening_outstanding
    for row in statement['rows']:
        amount = Decimal(str(row['amount']))
        running_outstanding = (running_outstanding + amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        values = [
            row['index'],
            format_cn_date(row['sales_date']),
            row['remark'],
            float(Decimal(str(row['received_weight'] or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)),
            float(Decimal(str(row['unit_price'] or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)),
            float(amount),
            '',
            float(running_outstanding),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_no, column=col, value=value)
            cell.border = border
            cell.alignment = center if col != 3 else left
            cell.font = text_font
            if col in {4, 5, 6, 7, 8} and value != '':
                cell.number_format = '#,##0.00'
        row_no += 1

    table_end_row = max(row_no, 13)
    while row_no <= table_end_row:
        for col in range(1, 9):
            cell = ws.cell(row=row_no, column=col)
            cell.border = border
            cell.alignment = center
        row_no += 1

    ws.merge_cells('A13:B13')
    ws.merge_cells('C13:H13')
    ws['A13'] = '本期客户累计未付款项'
    ws['A13'].alignment = left
    ws['C13'] = float(outstanding_amount)
    ws['C13'].number_format = '¥#,##0.00'
    ws['C13'].alignment = left

    ws.merge_cells('A14:B14')
    ws.merge_cells('C14:H14')
    ws['A14'] = '款项'
    ws['A14'].alignment = left
    ws['C14'] = f'RMB{statement["total_amount_uppercase"]}'
    ws['C14'].font = note_font
    ws['C14'].alignment = left

    for row in range(13, 15):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if row == 14:
                cell.border = border
            cell.font = note_font if row == 14 and col == 3 else text_font

    ws.merge_cells('A16:C16')
    ws.merge_cells('E16:H16')
    ws['A16'] = '数据核对无误，请确认回传，谢谢!'
    ws['E16'] = f'供货单位：{statement["seller_name"]}'
    ws['A16'].alignment = left
    ws['E16'].alignment = left

    ws.merge_cells('E17:H17')
    ws['A17'] = '客户确认（盖章 签名）'
    ws['E17'] = '联系电话：13342626882'
    ws['A17'].alignment = left
    ws['E17'].alignment = left

    ws.merge_cells('A18:C18')
    ws.merge_cells('E18:H18')
    cn_statement_date = format_cn_date(statement['statement_date'])
    ws['A18'] = f'日期：{cn_statement_date}'
    ws['E18'] = f'日期：{cn_statement_date}'
    ws['A18'].alignment = left
    ws['E18'].alignment = left

    for row in range(16, 19):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = text_font

    widths = {'A': 6, 'B': 14, 'C': 20, 'D': 11, 'E': 13, 'F': 17, 'G': 17, 'H': 17}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row, height in {1: 28, 2: 24, 3: 20, 4: 22, 5: 22, 13: 22, 14: 22, 16: 22, 17: 22, 18: 22}.items():
        ws.row_dimensions[row].height = height

    # Make Excel-to-PDF export fit a single A4 page width more reliably.
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_title_rows = '$1:$5'
    ws.print_area = f'$A$1:$H$18'
    wb.save(output_path)


def write_statement_files(statement: dict[str, Any], output_dir: Path, base_name: str) -> dict[str, str]:
    for suffix in ('.html', '.pdf', '.json'):
        stale_path = output_dir / f'{base_name}{suffix}'
        if stale_path.exists():
            stale_path.unlink()
    excel_path = output_dir / f'{base_name}.xlsx'
    render_excel(statement, excel_path)
    return {
        'excel': str(excel_path),
    }


def main():
    parser = argparse.ArgumentParser(
        description='Issue customer statement from the sales workbook.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Filters:
  --customer        Required. Customer full name or unique shorthand.
  --date-from       Optional sales date start.
  --date-to         Optional sales date end.
  --contract        Optional contract number.
  --paid            Optional: yes / no / all. Default: all.
  --statement-date  Optional statement date. Defaults to today.
  --multi-contract-mode  Optional: summary / split. Required when multiple contracts match.
        '''
    )
    parser.add_argument('--customer', required=True, help='Customer full name. Required.')
    parser.add_argument('--date-from', help='Sales date start, format YYYY.MM.DD.')
    parser.add_argument('--date-to', help='Sales date end, format YYYY.MM.DD.')
    parser.add_argument('--contract', help='Sales contract number.')
    parser.add_argument('--paid', default='all', choices=['yes', 'no', 'all'], help='Filter by paid status.')
    parser.add_argument('--statement-date', help='Statement date. Defaults to today.')
    parser.add_argument('--multi-contract-mode', choices=['summary', 'split'], help='How to handle multiple contracts.')
    args = parser.parse_args()

    statement_date = normalize_date_text(args.statement_date) or datetime.now().strftime(DATE_FMT)
    resolved_customer, match_mode, contracts = collect_matching_contracts(
        customer_input=args.customer,
        contract=args.contract,
        date_from=args.date_from,
        date_to=args.date_to,
        paid=args.paid,
    )

    if not contracts:
        print(json.dumps({
            'ok': False,
            'message': 'No matching contracts found for the provided filters.',
            'filters': {
                'customer': args.customer,
                'resolved_customer': resolved_customer,
                'match_mode': match_mode,
                'date_from': args.date_from,
                'date_to': args.date_to,
                'contract': args.contract,
                'paid': args.paid,
            },
        }, ensure_ascii=False, indent=2))
        raise SystemExit(1)

    if len(contracts) > 1 and not args.multi_contract_mode:
        print(json.dumps({
            'ok': False,
            'message': 'Multiple contracts matched. Please choose summary or split mode.',
            'resolved_customer': resolved_customer,
            'match_mode': match_mode,
            'matched_contracts': [
                {
                    'contract_no': c['contract_no'],
                    'sales_date': c['sales_date'],
                    'truck_count': c['truck_count'],
                    'paid': c['paid'],
                } for c in contracts
            ],
            'next_step': 'Rerun with --multi-contract-mode summary or --multi-contract-mode split',
        }, ensure_ascii=False, indent=2))
        raise SystemExit(1)

    output_dir = make_output_dir(resolved_customer, contracts)

    outputs = []
    file_stem = build_output_file_stem(resolved_customer, contracts)
    if args.multi_contract_mode == 'split':
        for contract in contracts:
            statement = build_statement_payload(
                customer_full_name=resolved_customer,
                statement_date=statement_date,
                contracts=[contract],
                mode='split',
            )
            single_stem = build_output_file_stem(resolved_customer, [contract])
            files = write_statement_files(statement, output_dir, single_stem)
            outputs.append({
                'contract_no': contract['contract_no'],
                'files': files,
            })
    else:
        statement = build_statement_payload(
            customer_full_name=resolved_customer,
            statement_date=statement_date,
            contracts=contracts,
            mode='summary' if len(contracts) > 1 else 'split',
        )
        files = write_statement_files(statement, output_dir, file_stem)
        outputs.append({
            'contract_no': statement['contract_display'],
            'files': files,
        })

    manifest = {
        'ok': True,
        'filters': {
            'customer': args.customer,
            'resolved_customer': resolved_customer,
            'match_mode': match_mode,
            'date_from': normalize_date_text(args.date_from),
            'date_to': normalize_date_text(args.date_to),
            'contract': args.contract,
            'paid': args.paid,
            'statement_date': statement_date,
            'multi_contract_mode': args.multi_contract_mode or ('split' if len(contracts) == 1 else None),
        },
        'matched_contracts': [
            {
                'contract_no': c['contract_no'],
                'sales_date': c['sales_date'],
                'truck_count': c['truck_count'],
                'paid': c['paid'],
            } for c in contracts
        ],
        'output_dir': str(output_dir),
        'outputs': outputs,
    }
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
