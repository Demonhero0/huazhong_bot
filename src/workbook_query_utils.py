#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import os
import re
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime
from pathlib import Path


REPO_ROOT = Path(__file__).resolve().parent.parent
ORDER_FILE = str(REPO_ROOT / '线材供应商提货明细龙虾版.xlsx')
SALES_FILE = str(REPO_ROOT / '线材客户送货明细龙虾版.xlsx')


def is_blank(value):
    return value is None or value == ''


def is_pending_receipt_cell(receipt_date, received_amount):
    # Some historical rows use 0 as a placeholder in V while no receipt has been recorded yet.
    return is_blank(receipt_date) and (is_blank(received_amount) or received_amount == 0)


def to_decimal(value, default='0'):
    if value is None or value == '':
        return Decimal(default)
    text = str(value).strip()
    if text.startswith('='):
        return Decimal(default)
    return Decimal(text).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def parse_date_text(value):
    if is_blank(value):
        return None
    text = str(value).strip()
    return datetime.strptime(text, '%Y.%m.%d')


def get_order_sheet_layout(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=4, column=col).value
        if is_blank(value):
            continue
        headers[str(value).strip()] = col

    layout = {
        'pickup_date': headers.get('提货日期', 8),
        'customer': headers.get('客户', 16),
        'truck_no': headers.get('车号', 9),
        'factory_weight': headers.get('出厂吨数', 10),
        'unit_price': headers.get('单价', 11),
        'pickup_amount': headers.get('提货金额', 12),
        'payment_date': headers.get('付款日期', 13),
        'payment_amount': headers.get('付款金额', 14),
        'balance': headers.get('余款', 15),
        # Supplier-side customer allocation fields are appended to the right in this workbook.
        'sell_price': 17,
        'delivery_mode': 18,
    }
    return layout


def extract_contract_price_from_title(value):
    text = str(value or '').strip().replace(' ', '')
    match = re.search(r'([0-9]+(?:\.[0-9]+)?)元', text)
    if not match:
        return None
    return float(match.group(1))


def load_order_workbook():
    return load_workbook(ORDER_FILE)


def load_sales_workbook():
    return load_workbook(SALES_FILE)


def get_order_sheet(wb, supplier):
    if supplier not in wb.sheetnames:
        raise KeyError(f'Supplier worksheet not found: {supplier}')
    return wb[supplier]


def get_sales_sheet(wb, customer):
    resolved_name, _ = resolve_sales_sheet_name(wb, customer)
    return wb[resolved_name]


def normalize_customer_name(name):
    text = str(name or '').strip()
    text = re.sub(r'[\s()（）\-.·,，]', '', text)
    for token in ['有限责任公司', '股份有限公司', '有限公司', '责任公司', '公司']:
        if text.endswith(token):
            text = text[:-len(token)]
    text = text.replace('省', '').replace('市', '')
    return text


def resolve_sales_sheet_name(wb, customer):
    """
    Resolve customer input to an existing sales worksheet name.
    Matching order:
    1. exact sheet name
    2. normalized unique match
    3. normalized containment unique match
    """
    if customer in wb.sheetnames:
        return customer, 'exact'

    normalized_input = normalize_customer_name(customer)
    if not normalized_input:
        raise KeyError('Customer worksheet not found: empty customer name')

    normalized_matches = []
    containment_matches = []
    for sheet_name in wb.sheetnames:
        normalized_sheet = normalize_customer_name(sheet_name)
        if normalized_sheet == normalized_input:
            normalized_matches.append(sheet_name)
        elif normalized_input in normalized_sheet or normalized_sheet in normalized_input:
            containment_matches.append(sheet_name)

    if len(normalized_matches) == 1:
        return normalized_matches[0], 'normalized'
    if len(normalized_matches) > 1:
        raise KeyError(
            f'Customer worksheet name "{customer}" is ambiguous. Candidates: {", ".join(normalized_matches)}'
        )

    if len(containment_matches) == 1:
        return containment_matches[0], 'containment'
    if len(containment_matches) > 1:
        raise KeyError(
            f'Customer worksheet name "{customer}" is ambiguous. Candidates: {", ".join(containment_matches)}'
        )

    raise KeyError(f'Customer worksheet not found: {customer}')


def iter_contract_rows(ws, contract_no):
    for row in range(5, ws.max_row + 1):
        if str(ws.cell(row=row, column=2).value) == str(contract_no):
            yield row


def get_contract_rows(ws, contract_no):
    return list(iter_contract_rows(ws, contract_no))


def resolve_order_contract_unit_price(ws, contract_no):
    layout = get_order_sheet_layout(ws)
    rows = get_contract_rows(ws, contract_no)
    if not rows:
        return None

    for row in rows:
        value = ws.cell(row=row, column=layout['unit_price']).value
        if not is_blank(value):
            return value

    for row in rows:
        parsed = extract_contract_price_from_title(ws.cell(row=row, column=1).value)
        if parsed is not None:
            return parsed

    return None


def resolve_sales_contract_sell_price(ws, contract_no):
    rows = get_contract_rows(ws, contract_no)
    if not rows:
        return None

    for row in rows:
        value = ws.cell(row=row, column=19).value
        if not is_blank(value):
            return value

    for row in rows:
        parsed = extract_contract_price_from_title(ws.cell(row=row, column=1).value)
        if parsed is not None:
            return parsed

    return None


def resolve_order_row_unit_price(ws, row):
    layout = get_order_sheet_layout(ws)
    contract_no = ws.cell(row=row, column=2).value
    if is_blank(contract_no):
        return None

    current = row
    fallback_unit_price = None
    while current >= 5 and str(ws.cell(row=current, column=2).value) == str(contract_no):
        parsed = extract_contract_price_from_title(ws.cell(row=current, column=1).value)
        if parsed is not None:
            return parsed
        unit_price = ws.cell(row=current, column=layout['unit_price']).value
        if fallback_unit_price is None and not is_blank(unit_price):
            fallback_unit_price = unit_price
        current -= 1

    return fallback_unit_price


def resolve_sales_row_sell_price(ws, row):
    contract_no = ws.cell(row=row, column=2).value
    if is_blank(contract_no):
        return None

    current = row
    while current >= 5 and str(ws.cell(row=current, column=2).value) == str(contract_no):
        sell_price = ws.cell(row=current, column=19).value
        if not is_blank(sell_price):
            return sell_price
        parsed = extract_contract_price_from_title(ws.cell(row=current, column=1).value)
        if parsed is not None:
            return parsed
        current -= 1

    return None


def get_contract_balance_snapshot(ws_values, contract_no):
    rows = get_contract_rows(ws_values, contract_no)
    if not rows:
        return None
    last_row = rows[-1]
    return {
        'row': last_row,
        'received_amount': to_decimal(ws_values.cell(row=last_row, column=22).value),
        'unreceived_amount': to_decimal(ws_values.cell(row=last_row, column=23).value),
    }


def resolve_sales_amount(ws_formulas, ws_values, row):
    cached_amount = ws_values.cell(row=row, column=20).value
    if not is_blank(cached_amount):
        return to_decimal(cached_amount)
    received_weight = to_decimal(ws_formulas.cell(row=row, column=18).value)
    unit_price = to_decimal(ws_formulas.cell(row=row, column=19).value)
    return (received_weight * unit_price).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def get_contract_receivable_snapshot(ws_formulas, ws_values, contract_no):
    rows = get_contract_rows(ws_formulas, contract_no)
    if not rows:
        return None

    total_sales_amount = Decimal('0.00')
    total_received_amount = Decimal('0.00')
    for row in rows:
        total_sales_amount += resolve_sales_amount(ws_formulas, ws_values, row)
        total_received_amount += to_decimal(ws_formulas.cell(row=row, column=22).value)

    last_row = rows[-1]
    outstanding_amount = (total_sales_amount - total_received_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    return {
        'row': last_row,
        'rows': rows,
        'total_sales_amount': total_sales_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'received_amount': total_received_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP),
        'unreceived_amount': outstanding_amount,
    }


def summarize_order_contracts(ws):
    layout = get_order_sheet_layout(ws)
    contracts = {}
    for row in range(5, ws.max_row + 1):
        contract_no = ws.cell(row=row, column=2).value
        if not contract_no:
            continue

        contract = contracts.setdefault(contract_no, {
            'contract_no': contract_no,
            'start_row': row,
            'end_row': row,
            'total_rows': 0,
            'transport': ws.cell(row=row, column=7).value,
            'brand': ws.cell(row=row, column=4).value,
            'spec': ws.cell(row=row, column=6).value,
            'unit_price': ws.cell(row=row, column=layout['unit_price']).value,
            'order_date': ws.cell(row=row, column=3).value,
            'pending_pickup_rows': [],
            'completed_pickup_rows': [],
            'pending_customer_rows': [],
            'assigned_customer_rows': [],
        })
        contract['end_row'] = row
        contract['total_rows'] += 1

        pickup_date = ws.cell(row=row, column=layout['pickup_date']).value
        truck_no = ws.cell(row=row, column=layout['truck_no']).value
        factory_weight = ws.cell(row=row, column=layout['factory_weight']).value
        customer = ws.cell(row=row, column=layout['customer']).value

        if is_blank(pickup_date) and is_blank(truck_no) and is_blank(factory_weight):
            contract['pending_pickup_rows'].append(row)
        else:
            contract['completed_pickup_rows'].append(row)

        if is_blank(customer):
            contract['pending_customer_rows'].append(row)
        else:
            contract['assigned_customer_rows'].append(row)

    result = []
    for contract in contracts.values():
        result.append({
            'contract_no': contract['contract_no'],
            'start_row': contract['start_row'],
            'end_row': contract['end_row'],
            'total_rows': contract['total_rows'],
            'transport': contract['transport'],
            'brand': contract['brand'],
            'spec': contract['spec'],
            'unit_price': contract['unit_price'],
            'order_date': contract['order_date'],
            'pending_pickup_rows': contract['pending_pickup_rows'],
            'completed_pickup_rows': contract['completed_pickup_rows'],
            'pending_customer_rows': contract['pending_customer_rows'],
            'assigned_customer_rows': contract['assigned_customer_rows'],
            'pending_pickup_count': len(contract['pending_pickup_rows']),
            'completed_pickup_count': len(contract['completed_pickup_rows']),
            'pending_customer_count': len(contract['pending_customer_rows']),
            'assigned_customer_count': len(contract['assigned_customer_rows']),
        })
    return sorted(result, key=lambda item: str(item['contract_no']))


def summarize_sales_contracts(ws):
    contracts = {}
    for row in range(5, ws.max_row + 1):
        contract_no = ws.cell(row=row, column=2).value
        if not contract_no:
            continue

        contract = contracts.setdefault(contract_no, {
            'contract_no': contract_no,
            'start_row': row,
            'end_row': row,
            'total_rows': 0,
            'supplier': ws.cell(row=row, column=12).value,
            'brand': ws.cell(row=row, column=4).value,
            'spec': ws.cell(row=row, column=5).value,
            'sales_date': ws.cell(row=row, column=3).value,
            'sell_price': ws.cell(row=row, column=19).value,
            'pending_delivery_rows': [],
            'completed_delivery_rows': [],
        })
        contract['end_row'] = row
        contract['total_rows'] += 1

        delivery_date = ws.cell(row=row, column=8).value
        truck_no = ws.cell(row=row, column=16).value
        factory_weight = ws.cell(row=row, column=17).value
        received_weight = ws.cell(row=row, column=18).value

        if is_blank(delivery_date) and is_blank(truck_no) and is_blank(factory_weight) and is_blank(received_weight):
            contract['pending_delivery_rows'].append(row)
        else:
            contract['completed_delivery_rows'].append(row)

    result = []
    for contract in contracts.values():
        result.append({
            'contract_no': contract['contract_no'],
            'start_row': contract['start_row'],
            'end_row': contract['end_row'],
            'total_rows': contract['total_rows'],
            'supplier': contract['supplier'],
            'brand': contract['brand'],
            'spec': contract['spec'],
            'sales_date': contract['sales_date'],
            'sell_price': contract['sell_price'],
            'pending_delivery_rows': contract['pending_delivery_rows'],
            'completed_delivery_rows': contract['completed_delivery_rows'],
            'pending_delivery_count': len(contract['pending_delivery_rows']),
            'completed_delivery_count': len(contract['completed_delivery_rows']),
        })
    return sorted(result, key=lambda item: str(item['contract_no']))


def find_pending_order_rows(ws, contract_no):
    layout = get_order_sheet_layout(ws)
    rows = []
    for row in iter_contract_rows(ws, contract_no):
        pickup_date = ws.cell(row=row, column=layout['pickup_date']).value
        truck_no = ws.cell(row=row, column=layout['truck_no']).value
        factory_weight = ws.cell(row=row, column=layout['factory_weight']).value
        if not (is_blank(pickup_date) and is_blank(truck_no) and is_blank(factory_weight)):
            continue
        rows.append({
            'row': row,
            'brand': ws.cell(row=row, column=4).value,
            'spec': ws.cell(row=row, column=6).value,
            'transport': ws.cell(row=row, column=7).value,
            'customer': ws.cell(row=row, column=layout['customer']).value,
            'sell_price': ws.cell(row=row, column=layout['sell_price']).value,
            'delivery_mode': ws.cell(row=row, column=layout['delivery_mode']).value,
            'dock': ws.cell(row=row, column=5).value,
            'pickup_date': pickup_date,
            'truck_no': truck_no,
            'factory_weight': factory_weight,
        })
    return rows


def find_pending_sales_rows(ws, contract_no):
    rows = []
    for row in iter_contract_rows(ws, contract_no):
        delivery_date = ws.cell(row=row, column=8).value
        truck_no = ws.cell(row=row, column=16).value
        factory_weight = ws.cell(row=row, column=17).value
        received_weight = ws.cell(row=row, column=18).value
        if not (is_blank(delivery_date) and is_blank(truck_no) and is_blank(factory_weight) and is_blank(received_weight)):
            continue
        rows.append({
            'row': row,
            'supplier': ws.cell(row=row, column=12).value,
            'brand': ws.cell(row=row, column=4).value,
            'spec': ws.cell(row=row, column=5).value,
            'transport_mode': ws.cell(row=row, column=14).value,
            'delivery_mode': ws.cell(row=row, column=14).value,
            'sell_price': ws.cell(row=row, column=19).value,
            'delivery_date': delivery_date,
            'truck_no': truck_no,
            'factory_weight': factory_weight,
            'received_weight': received_weight,
        })
    return rows


def summarize_receivable_contracts(ws_formulas, ws_values, date_from=None, date_to=None, settled='no', contract_no=None):
    contracts = {}
    for row in range(5, ws_formulas.max_row + 1):
        current_contract = ws_formulas.cell(row=row, column=2).value
        if is_blank(current_contract):
            continue
        current_contract = str(current_contract)
        if contract_no and current_contract != str(contract_no):
            continue

        sales_date = ws_formulas.cell(row=row, column=3).value
        sales_dt = parse_date_text(sales_date) if not is_blank(sales_date) else None
        if date_from and sales_dt and sales_dt < date_from:
            continue
        if date_to and sales_dt and sales_dt > date_to:
            continue

        contract = contracts.setdefault(current_contract, {
            'contract_no': current_contract,
            'sales_date': sales_date,
            'start_row': row,
            'end_row': row,
            'total_rows': 0,
            'supplier': ws_formulas.cell(row=row, column=12).value,
            'brand': ws_formulas.cell(row=row, column=4).value,
            'spec': ws_formulas.cell(row=row, column=5).value,
            'sell_price': ws_formulas.cell(row=row, column=19).value,
            'receipt_filled_rows': [],
            'receipt_pending_rows': [],
        })
        contract['end_row'] = row
        contract['total_rows'] += 1

        receipt_date = ws_formulas.cell(row=row, column=21).value
        received_amount = ws_formulas.cell(row=row, column=22).value
        if is_pending_receipt_cell(receipt_date, received_amount):
            contract['receipt_pending_rows'].append(row)
        else:
            contract['receipt_filled_rows'].append(row)

    result = []
    for contract in contracts.values():
        balance = get_contract_receivable_snapshot(ws_formulas, ws_values, contract['contract_no'])
        current_received_amount = balance['received_amount'] if balance else Decimal('0.00')
        current_unreceived_amount = balance['unreceived_amount'] if balance else Decimal('0.00')
        is_settled = current_unreceived_amount == Decimal('0.00')
        if settled == 'yes' and not is_settled:
            continue
        if settled == 'no' and is_settled:
            continue
        result.append({
            'contract_no': contract['contract_no'],
            'sales_date': contract['sales_date'],
            'start_row': contract['start_row'],
            'end_row': contract['end_row'],
            'total_rows': contract['total_rows'],
            'supplier': contract['supplier'],
            'brand': contract['brand'],
            'spec': contract['spec'],
            'sell_price': contract['sell_price'],
            'receipt_filled_rows': contract['receipt_filled_rows'],
            'receipt_pending_rows': contract['receipt_pending_rows'],
            'receipt_filled_count': len(contract['receipt_filled_rows']),
            'receipt_pending_count': len(contract['receipt_pending_rows']),
            'balance_row': balance['row'] if balance else None,
            'current_received_amount': f'{current_received_amount:.2f}',
            'current_unreceived_amount': f'{current_unreceived_amount:.2f}',
            'is_settled': is_settled,
        })
    return sorted(result, key=lambda item: str(item['contract_no']))


def find_first_receipt_pending_row(ws, contract_no):
    for row in iter_contract_rows(ws, contract_no):
        receipt_date = ws.cell(row=row, column=21).value
        received_amount = ws.cell(row=row, column=22).value
        if is_pending_receipt_cell(receipt_date, received_amount):
            return row
    return None


def summarize_invoice_rows(ws_formulas, ws_values, customer_name, date_from=None, date_to=None, invoiced='no', contract_no=None):
    rows = []
    for row in range(5, ws_formulas.max_row + 1):
        current_contract_no = ws_formulas.cell(row=row, column=2).value
        if is_blank(current_contract_no):
            continue
        if contract_no and str(current_contract_no) != str(contract_no):
            continue

        delivery_date = ws_formulas.cell(row=row, column=8).value
        if is_blank(delivery_date):
            continue

        delivery_dt = parse_date_text(delivery_date)
        if date_from and delivery_dt < date_from:
            continue
        if date_to and delivery_dt > date_to:
            continue

        invoice_time = ws_formulas.cell(row=row, column=24).value
        is_invoiced = not is_blank(invoice_time)
        if invoiced == 'yes' and not is_invoiced:
            continue
        if invoiced == 'no' and is_invoiced:
            continue

        sales_amount = resolve_sales_amount(ws_formulas, ws_values, row)
        rows.append({
            'row': row,
            'customer': customer_name,
            'contract_no': str(current_contract_no),
            'sales_date': ws_formulas.cell(row=row, column=3).value,
            'delivery_date': delivery_date,
            'supplier': ws_formulas.cell(row=row, column=12).value,
            'brand': ws_formulas.cell(row=row, column=4).value,
            'spec': ws_formulas.cell(row=row, column=5).value,
            'truck_no': ws_formulas.cell(row=row, column=16).value,
            'received_weight': ws_formulas.cell(row=row, column=18).value,
            'unit_price': ws_formulas.cell(row=row, column=19).value,
            'sales_amount': f'{sales_amount:.2f}',
            'receipt_date': ws_formulas.cell(row=row, column=21).value,
            'received_amount': f'{to_decimal(ws_formulas.cell(row=row, column=22).value):.2f}',
            'invoice_time': invoice_time,
            'is_invoiced': is_invoiced,
        })
    return rows


def find_pending_invoice_rows(ws, contract_no):
    rows = []
    for row in iter_contract_rows(ws, contract_no):
        delivery_date = ws.cell(row=row, column=8).value
        invoice_time = ws.cell(row=row, column=24).value
        if is_blank(delivery_date) or not is_blank(invoice_time):
            continue
        rows.append({
            'row': row,
            'delivery_date': delivery_date,
            'truck_no': ws.cell(row=row, column=16).value,
            'received_weight': ws.cell(row=row, column=18).value,
            'unit_price': ws.cell(row=row, column=19).value,
            'receipt_date': ws.cell(row=row, column=21).value,
            'received_amount': ws.cell(row=row, column=22).value,
            'invoice_time': invoice_time,
        })
    return rows
