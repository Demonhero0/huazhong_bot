#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Steel Wire Sales Entry Script
Fill customer info into supplier order file AND sales file

Usage:
  # List available contracts
  python3 sales_entry.py -s <supplier>
  
  # Single customer (trucks defaults to 1)
  python3 sales_entry.py -s <supplier> -n <contract_no> -c <customer> -p <price> -d <delivery> [-t <trucks>] [--benchmark <benchmark>] [--price-diff <price_diff>] [--sales-date <sales_date>]
  
  # Multiple customers
  python3 sales_entry.py -s <supplier> -n <contract_no> -c <customer1> -p <price1> -d <delivery1> -t <trucks1> [--benchmark <benchmark1>] [--price-diff <price_diff1>] -c <customer2> -p <price2> -d <delivery2> [--benchmark <benchmark2>] [--price-diff <price_diff2>] [--sales-date <sales_date>]
"""

from openpyxl import load_workbook
from datetime import datetime
import sys
import os
import argparse

from workbook_query_utils import ORDER_FILE, SALES_FILE, resolve_sales_sheet_name


def find_contracts_with_empty_cells(ws, supplier=None):
    """
    Find all contracts with empty customer cells in the worksheet
    Returns: list of (contract_no, start_row, end_row, empty_count, transport_method)
    """
    contracts = {}
    
    for row in range(5, ws.max_row + 1):
        contract_no = ws.cell(row=row, column=2).value  # Column B
        if not contract_no:
            continue
        
        customer = ws.cell(row=row, column=16).value  # Column P
        transport = ws.cell(row=row, column=7).value  # Column G
        
        if contract_no not in contracts:
            contracts[contract_no] = {
                'start_row': row,
                'end_row': row,
                'empty_rows': [],
                'transport': transport
            }
        
        contracts[contract_no]['end_row'] = row
        
        if customer is None or customer == '':
            contracts[contract_no]['empty_rows'].append(row)
    
    result = []
    for contract_no, data in contracts.items():
        empty_count = len(data['empty_rows'])
        if empty_count > 0:
            result.append((contract_no, data['start_row'], data['end_row'], empty_count, data['transport']))
    
    return result


def find_empty_row_in_sales(ws):
    """
    Find the first completely empty row in sales worksheet
    Returns: row number
    """
    for row in range(5, ws.max_row + 100):  # Check up to 100 rows beyond current max
        has_data = False
        for col in range(1, 25):  # Check first 24 columns
            val = ws.cell(row=row, column=col).value
            if val is not None and val != '':
                has_data = True
                break
        if not has_data:
            return row
    return ws.max_row + 1


def read_order_info(ws, row):
    """
    Read order information from a specific row
    Returns: dict with order info
    """
    return {
        'brand': ws.cell(row=row, column=4).value,  # D 列：品牌
        'spec': ws.cell(row=row, column=6).value,   # F 列：规格
        'order_price': ws.cell(row=row, column=11).value,  # K 列：单价
        'transport': ws.cell(row=row, column=7).value,     # G 列：运输方式
        'order_date': ws.cell(row=row, column=3).value,    # C 列：订货日期
    }


def enter_sales(supplier, contract_no, customers, prices, deliveries, trucks_list, benchmarks=None, price_diffs=None, sales_date=None):
    """
    Enter sales information to supplier order file AND sales file
    
    Args:
    - supplier: Worksheet name (e.g., "浙江凯航")
    - contract_no: Contract number to match (e.g., "2026031801"), or None to list all
    - customers: List of customer names
    - prices: List of prices
    - deliveries: List of delivery types
    - trucks_list: List of truck counts (default 1 for each)
    - benchmarks: List of benchmark strings for F column, optional
    - price_diffs: List of price difference strings for G column, optional
    - sales_date: Sales date (e.g., "2026.03.20"), optional, defaults to today
    """
    
    # ========== PART 1: Fill Order File ==========
    print(f'\n=== Part 1: Fill Order File ===')
    if benchmarks is None:
        benchmarks = []
    if price_diffs is None:
        price_diffs = []
    if not sales_date:
        sales_date = datetime.now().strftime('%Y.%m.%d')
    contract_no_sales = sales_date.replace('.', '') + '01'
    
    # Open order file
    if not os.path.exists(ORDER_FILE):
        print(f'\n❌ Error: Order file not found: {ORDER_FILE}')
        return False
    
    wb_order = load_workbook(ORDER_FILE)
    wb_sales = load_workbook(SALES_FILE) if os.path.exists(SALES_FILE) else None
    
    if supplier not in wb_order.sheetnames:
        print(f'\n❌ Error: Worksheet "{supplier}" not found in order file!')
        print(f'Available worksheets: {", ".join(wb_order.sheetnames)}')
        return False
    
    ws_order = wb_order[supplier]
    
    # Find all contracts with empty customer cells
    all_contracts = find_contracts_with_empty_cells(ws_order, supplier)
    
    if not all_contracts:
        print(f'\n❌ Error: No contracts with empty customer cells found in "{supplier}"!')
        return False
    
    # If no contract_no provided, list available contracts and ask user to confirm
    if not contract_no:
        print(f'\n📋 Available contracts with empty customer cells in "{supplier}":')
        print('=' * 100)
        print(f'{"Contract No":<15} {"Rows":<10} {"Empty Cells":<12} {"Transport":<10}')
        print('-' * 100)
        for c_no, start, end, empty_count, transport in sorted(all_contracts, key=lambda x: x[0]):
            print(f'{c_no:<15} {start}-{end:<9} {empty_count:<12} {transport}')
        print('-' * 100)
        print(f'\n⚠️  Please specify which contract to use with -n parameter.')
        print(f'\n💡 Usage example:')
        print(f'  python3 sales_entry.py -s "{supplier}" -n 2026032401 -c "东莞建安" -p 3400 -t 2')
        print(f'\n  (送货方式默认为"送到"，如需"自提"请添加 -d 自提)')
        return True
    
    # Find the specific contract
    contract_data = None
    for c_no, start, end, empty_count, transport in all_contracts:
        if c_no == contract_no:
            contract_data = (c_no, start, end, empty_count, transport)
            break
    
    if not contract_data:
        print(f'\n❌ Error: Contract No "{contract_no}" not found or has no empty customer cells!')
        print(f'\n📋 Available contracts in "{supplier}":')
        print('=' * 100)
        for c_no, start, end, empty_count, transport in sorted(all_contracts, key=lambda x: x[0]):
            print(f'{c_no:<15} {start}-{end:<9} {empty_count:<12} {transport}')
        return False
    
    c_no, start_row, end_row, empty_count, transport = contract_data
    empty_rows = []
    for row in range(start_row, end_row + 1):
        customer = ws_order.cell(row=row, column=16).value
        if customer is None or customer == '':
            empty_rows.append(row)
    
    print(f'\nContract No: {c_no}')
    print(f'Contract rows: {start_row}-{end_row}')
    print(f'Empty customer cells: {empty_count} (rows: {empty_rows})')
    print(f'Transport method: {transport}')
    
    # Validate input arrays have same length
    num_customers = len(customers)
    if len(prices) != num_customers or len(deliveries) != num_customers:
        print(f'\n❌ Error: Number of -c, -p, and -d arguments must be the same!')
        return False
    
    # Extend trucks_list and deliveries to match num_customers (defaults: trucks=1, delivery="送到")
    while len(trucks_list) < num_customers:
        trucks_list.append(1)
    while len(deliveries) < num_customers:
        deliveries.append('送到')  # Default delivery type is "送到"
    while len(benchmarks) < num_customers:
        benchmarks.append('')
    while len(price_diffs) < num_customers:
        price_diffs.append('')
    
    # Build customer info list
    parsed_customers = []
    for i in range(num_customers):
        input_customer = customers[i]
        resolved_customer = input_customer
        customer_match_mode = 'new'
        if wb_sales is not None:
            try:
                resolved_customer, customer_match_mode = resolve_sales_sheet_name(wb_sales, input_customer)
            except KeyError:
                resolved_customer = input_customer
                customer_match_mode = 'new'
        parsed_customers.append({
            'customer': resolved_customer,
            'input_customer': input_customer,
            'price': prices[i],
            'delivery': deliveries[i] if deliveries[i] else '送到',  # Default to "送到"
            'trucks': trucks_list[i] if i < len(trucks_list) else 1,
            'benchmark': benchmarks[i] if i < len(benchmarks) else '',
            'price_diff': price_diffs[i] if i < len(price_diffs) else '',
            'customer_match_mode': customer_match_mode,
        })
    
    # Validate delivery types
    for i, c in enumerate(parsed_customers, 1):
        if c['delivery'] not in ['自提', '送到']:
            print(f'\n❌ Error: Invalid delivery type for customer #{i}: "{c["delivery"]}"')
            return False
    
    # Validate truck counts
    for i, c in enumerate(parsed_customers, 1):
        if c['trucks'] < 1:
            print(f'\n❌ Error: Invalid truck count for customer #{i}: {c["trucks"]}')
            return False
    
    # Calculate total trucks needed
    total_trucks = sum(c['trucks'] for c in parsed_customers)
    
    print(f'\nCustomers to fill ({num_customers}):')
    for i, c in enumerate(parsed_customers, 1):
        line = f'  {i}. {c["customer"]}, {c["price"]} yuan, {c["delivery"]}, {c["trucks"]} truck(s)'
        if c['input_customer'] != c['customer']:
            line += f' [from "{c["input_customer"]}" via {c["customer_match_mode"]} match]'
        print(line)
    print(f'\nTotal trucks needed: {total_trucks}')
    print(f'Available empty rows: {empty_count}')
    
    # Check if enough empty rows
    if total_trucks > empty_count:
        print(f'\n❌ Error: Not enough empty rows in order file!')
        print(f'   Needed: {total_trucks} rows')
        print(f'   Available: {empty_count} rows')
        print(f'   Shortage: {total_trucks - empty_count} rows')
        return False
    
    # Fill in the order file and collect order info for sales file
    current_empty_idx = 0
    order_info_list = []  # Store order info for each customer
    
    for cust in parsed_customers:
        # Read order info from the first row for this customer
        first_row = empty_rows[current_empty_idx]
        order_info = read_order_info(ws_order, first_row)
        order_info['customer'] = cust['customer']
        order_info['sell_price'] = cust['price']
        order_info['delivery'] = cust['delivery']
        order_info['trucks'] = cust['trucks']
        order_info['contract_no'] = c_no
        order_info['benchmark'] = cust['benchmark']
        order_info['price_diff'] = cust['price_diff']
        order_info_list.append(order_info)
        
        for _ in range(cust['trucks']):
            target_row = empty_rows[current_empty_idx]
            
            # Column P (16): Customer
            ws_order.cell(row=target_row, column=16, value=cust['customer'])
            # Column Q (17): Sell Price
            ws_order.cell(row=target_row, column=17, value=cust['price'])
            # Column R (18): Delivery Type
            ws_order.cell(row=target_row, column=18, value=cust['delivery'])
            
            print(f'  ✅ Order Row {target_row}: {cust["customer"]}, {cust["price"]}, {cust["delivery"]}')
            current_empty_idx += 1
    
    # Save order file
    wb_order.save(ORDER_FILE)
    print(f'\n✅ Order file updated: {ORDER_FILE}')
    
    # ========== PART 2: Fill Sales File ==========
    print(f'\n=== Part 2: Fill Sales File ===')
    
    if not os.path.exists(SALES_FILE):
        print(f'\n❌ Error: Sales file not found: {SALES_FILE}')
        return False

    # Process each customer
    for order_info in order_info_list:
        customer_name = order_info['customer']
        trucks = order_info['trucks']
        
        # Find or create customer worksheet
        if customer_name not in wb_sales.sheetnames:
            print(f'\n⚠️  Warning: Customer worksheet "{customer_name}" not found, creating...')
            wb_sales.create_sheet(customer_name)
        
        ws_sales = wb_sales[customer_name]
        
        # Find first completely empty row (start row for this sales entry)
        sales_start_row = find_empty_row_in_sales(ws_sales)
        print(f'\nCustomer: {customer_name}')
        print(f'Sales start row: {sales_start_row}')
        print(f'Trucks: {trucks}')
        print(f'Order info: Brand={order_info["brand"]}, Spec={order_info["spec"]}, Price={order_info["sell_price"]}')
        print(f'Sales Date: {sales_date}')
        if order_info.get('benchmark'):
            print(f'Benchmark: {order_info["benchmark"]}')
        if order_info.get('price_diff'):
            print(f'Price Diff: {order_info["price_diff"]}')
        
        # Generate sales contract info
        sales_contract = f'{int(order_info["sell_price"]) }元{trucks}车'
        
        order_price = order_info.get('order_price', 0)
        sell_price = order_info['sell_price']
        
        # Fill sales information for each truck (one row per truck)
        for truck_idx in range(trucks):
            current_row = sales_start_row + truck_idx
            prev_row = current_row - 1
            
            # A 列：销售合同（只在第一行填写）
            if truck_idx == 0:
                ws_sales.cell(row=current_row, column=1, value=sales_contract)
            # B 列：合同编号（每行都填）
            ws_sales.cell(row=current_row, column=2, value=contract_no_sales)
            # C 列：销售日期（每行都填）
            ws_sales.cell(row=current_row, column=3, value=sales_date)
            # D 列：品牌
            ws_sales.cell(row=current_row, column=4, value=order_info['brand'])
            # E 列：规格
            ws_sales.cell(row=current_row, column=5, value=order_info['spec'])
            # F 列：对标
            if order_info.get('benchmark'):
                ws_sales.cell(row=current_row, column=6, value=order_info['benchmark'])
            # G 列：成交价差
            if order_info.get('price_diff'):
                ws_sales.cell(row=current_row, column=7, value=order_info['price_diff'])
            # H 列：送货日期（不填，跳过）
            # L 列：供应商
            ws_sales.cell(row=current_row, column=12, value=supplier)
            # M 列：提货价
            ws_sales.cell(row=current_row, column=13, value=order_price)
            # N 列：自提/送到
            ws_sales.cell(row=current_row, column=14, value=order_info['delivery'])
            # S 列：单价
            ws_sales.cell(row=current_row, column=19, value=sell_price)
            # T 列：销售金额（公式，每行都有）
            ws_sales.cell(row=current_row, column=20, value=f'=R{current_row}*S{current_row}')
            # W 列：未收款金额（累计余额公式）
            # 公式逻辑：W{current} = W{prev} - T{current} + V{current}
            # W 为负数表示应收（欠款），0 表示结清，正数表示多收
            if truck_idx == 0:
                # 检查上一行（prev_row）是否有 W 列数据
                prev_w = ws_sales.cell(row=prev_row, column=23).value
                if prev_w and (str(prev_w).startswith('=') or (isinstance(prev_w, (int, float)) and prev_w != 0)):
                    # 上一行有公式或数值，累加
                    ws_sales.cell(row=current_row, column=23, value=f'=W{prev_row}-T{current_row}+V{current_row}')
                else:
                    # 第一笔或上一行 W 为空/0：W = -T + V
                    # V 为空时，W = -T（负数表示应收/欠款）
                    ws_sales.cell(row=current_row, column=23, value=f'=-T{current_row}+V{current_row}')
            else:
                # 非第一行，累加上一行的 W
                ws_sales.cell(row=current_row, column=23, value=f'=W{prev_row}-T{current_row}+V{current_row}')
            
            print(f'  ✅ Sales Row {current_row}:')
            if truck_idx == 0:
                print(f'     Contract: {sales_contract}')
            print(f'     Brand: {order_info["brand"]}, Spec: {order_info["spec"]}')
            print(f'     Order Price: {order_price}, Sell Price: {sell_price}')
            print(f'     Supplier: {supplier}, Delivery: {order_info["delivery"]}')
    
    # Save sales file
    wb_sales.save(SALES_FILE)
    print(f'\n✅ Sales file updated: {SALES_FILE}')
    print(f'\n✅ All entries completed!')
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Steel Wire Sales Entry Script',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  # List available contracts
  python3 sales_entry.py -s "浙江凯航"
  
  # Single customer (trucks defaults to 1)
  python3 sales_entry.py -s "浙江凯航" -n 2026032301 -c "东莞建安" -p 3400 -d 自提

  # Single customer with benchmark and price diff
  python3 sales_entry.py -s "浙江凯航" -n 2026032301 -c "东莞建安" -p 3400 -d 自提 --benchmark "迁安自提" --price-diff 80

  # Single customer with explicit sales date
  python3 sales_entry.py -s "浙江凯航" -n 2026032301 -c "东莞建安" -p 3400 -d 自提 --sales-date 2026.03.20
  
  # Single customer with explicit truck count
  python3 sales_entry.py -s "浙江凯航" -n 2026032301 -c "东莞建安" -p 3400 -d 自提 -t 2
  
  # Multiple customers
  python3 sales_entry.py -s "浙江凯航" -n 2026032301 -c "东莞建安" -p 3400 -d 自提 -t 2 -c "中山富华" -p 3420 -d 送到
        '''
    )
    
    # Required arguments
    parser.add_argument('-s', '--supplier', required=True, help='Worksheet name (e.g., "浙江凯航")')
    parser.add_argument('-n', '--contract', help='Contract number to match (e.g., "2026032301"). If omitted, lists available contracts.')
    
    # Customer info arguments (can be repeated for multiple customers)
    parser.add_argument('-c', '--customer', action='append', default=[], help='Customer name (e.g., "东莞建安"). Can be used multiple times.')
    parser.add_argument('-p', '--price', type=float, action='append', default=[], help='Selling price (e.g., 3400). Can be used multiple times.')
    parser.add_argument('-d', '--delivery', action='append', default=[], help='Delivery type: 自提 or 送到 (e.g., "送到"). Can be used multiple times.')
    parser.add_argument('-t', '--trucks', type=int, action='append', default=[], help='Number of trucks (default: 1). Can be used multiple times.')
    parser.add_argument('--benchmark', action='append', default=[], help='Benchmark / 对标 for sales file column F. Can be used multiple times.')
    parser.add_argument('--price-diff', action='append', default=[], help='Price difference / 成交价差 for sales file column G. Can be used multiple times.')
    parser.add_argument('--sales-date', help='Sales date (e.g., "2026.03.20"). Defaults to today if omitted.')
    
    args = parser.parse_args()
    
    # If no customer info provided
    if not args.customer and not args.contract:
        # Just list available contracts
        enter_sales(args.supplier, None, [], [], [], [])
    elif not args.customer:
        print('\n❌ Error: No customer info provided!')
        print('Use -c, -p, -d to specify customer info (trucks -t is optional, defaults to 1)')
        print('\nExample:')
        print(f'  python3 sales_entry.py -s "{args.supplier}" -n 2026032301 -c "东莞建安" -p 3400 -d 自提')
        sys.exit(1)
    else:
        # Execute entry
        enter_sales(
            supplier=args.supplier,
            contract_no=args.contract,
            customers=args.customer,
            prices=args.price,
            deliveries=args.delivery,
            trucks_list=args.trucks,
            benchmarks=args.benchmark,
            price_diffs=args.price_diff,
            sales_date=args.sales_date
        )
