#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Steel Wire Invoice Entry Script
Fill invoice time into the customer sales workbook.

Usage:
  python3 src/invoice_entry.py -c <customer> -n <sales_contract> [--invoice-date <YYYY.MM.DD>] [--count <n>] [--row <row> ...]
  python3 src/invoice_entry.py [-c <customer>] --date-from <YYYY.MM.DD> --date-to <YYYY.MM.DD> [--invoice-date <YYYY.MM.DD>]
"""

import argparse
import os
import sys
from datetime import datetime

from openpyxl import load_workbook

from workbook_query_utils import SALES_FILE, find_pending_invoice_rows, parse_date_text, resolve_sales_sheet_name, summarize_invoice_rows


DATE_FMT = '%Y.%m.%d'


def enter_invoice(customer, contract_no, invoice_date=None, count=None, rows=None):
    if not invoice_date:
        invoice_date = datetime.now().strftime(DATE_FMT)

    if count is not None and count <= 0:
        print('\n❌ Error: Invoice count must be greater than 0!')
        return False

    if not os.path.exists(SALES_FILE):
        print(f'\n❌ Error: Sales file not found: {SALES_FILE}')
        return False

    wb = load_workbook(SALES_FILE)
    try:
        resolved_customer, match_mode = resolve_sales_sheet_name(wb, customer)
    except KeyError:
        print(f'\n❌ Error: Customer worksheet "{customer}" not found in sales file!')
        print(f'Available worksheets: {", ".join(wb.sheetnames)}')
        return False

    ws = wb[resolved_customer]
    pending_rows = find_pending_invoice_rows(ws, contract_no)
    if not pending_rows:
        print(f'\n❌ Error: No pending invoice rows found for customer "{customer}" contract "{contract_no}"!')
        print('The contract may not exist, may have no delivered rows yet, or may already be fully invoiced.')
        return False

    pending_row_numbers = [item['row'] for item in pending_rows]
    if rows:
        target_rows = []
        invalid_rows = []
        for row in rows:
            if row not in pending_row_numbers:
                invalid_rows.append(row)
            else:
                target_rows.append(row)
        if invalid_rows:
            print('\n❌ Error: Some specified rows are not pending invoice rows for this contract!')
            print(f'   Invalid rows: {invalid_rows}')
            print(f'   Pending rows: {pending_row_numbers}')
            return False
    elif count is not None:
        if count > len(pending_rows):
            print('\n❌ Error: Invoice count exceeds pending invoice rows!')
            print(f'   Requested count: {count}')
            print(f'   Pending rows: {len(pending_rows)}')
            return False
        target_rows = pending_row_numbers[:count]
    else:
        target_rows = pending_row_numbers

    for row in target_rows:
        ws.cell(row=row, column=24, value=invoice_date)

    wb.save(SALES_FILE)

    print('\n=== Invoice Entry ===')
    print(f'Customer: {customer}')
    if resolved_customer != customer:
        print(f'Resolved Customer Sheet: {resolved_customer} ({match_mode})')
    print(f'Sales Contract: {contract_no}')
    print(f'Invoice Date: {invoice_date}')
    print(f'Rows Invoiced: {target_rows}')
    print(f'Pending Rows Before Write: {pending_row_numbers}')
    print(f'\n✅ Sales file updated: {SALES_FILE}')
    print('✅ Invoice entry completed!')
    return True


def enter_invoice_by_delivery_range(customer, date_from_text, date_to_text, invoice_date=None):
    if not invoice_date:
        invoice_date = datetime.now().strftime(DATE_FMT)

    if not os.path.exists(SALES_FILE):
        print(f'\n❌ Error: Sales file not found: {SALES_FILE}')
        return False

    date_from = parse_date_text(date_from_text)
    date_to = parse_date_text(date_to_text)
    if date_from is None or date_to is None:
        print('\n❌ Error: Both --date-from and --date-to are required in YYYY.MM.DD format!')
        return False
    if date_from > date_to:
        print('\n❌ Error: --date-from cannot be later than --date-to!')
        return False

    wb = load_workbook(SALES_FILE)
    wb_values = load_workbook(SALES_FILE, data_only=True)

    try:
        if customer:
            resolved_customer, match_mode = resolve_sales_sheet_name(wb, customer)
            customer_names = [resolved_customer]
        else:
            resolved_customer = None
            match_mode = None
            customer_names = list(wb.sheetnames)
    except KeyError:
        print(f'\n❌ Error: Customer worksheet "{customer}" not found in sales file!')
        print(f'Available worksheets: {", ".join(wb.sheetnames)}')
        return False

    target_rows = []
    for customer_name in customer_names:
        invoice_rows = summarize_invoice_rows(
            ws_formulas=wb[customer_name],
            ws_values=wb_values[customer_name],
            customer_name=customer_name,
            date_from=date_from,
            date_to=date_to,
            invoiced='no',
        )
        for item in invoice_rows:
            target_rows.append((customer_name, item['row'], item['contract_no']))

    if not target_rows:
        scope = f'customer "{customer}" ' if customer else ''
        print(f'\n❌ Error: No pending invoice rows found for {scope}delivery dates {date_from_text} to {date_to_text}!')
        return False

    for customer_name, row, _contract_no in target_rows:
        wb[customer_name].cell(row=row, column=24, value=invoice_date)

    wb.save(SALES_FILE)

    print('\n=== Invoice Entry By Delivery Date Range ===')
    if customer:
        print(f'Customer: {customer}')
        if resolved_customer != customer:
            print(f'Resolved Customer Sheet: {resolved_customer} ({match_mode})')
    else:
        print('Customer Scope: all customer sheets')
    print(f'Delivery Date Range: {date_from_text} ~ {date_to_text}')
    print(f'Invoice Date: {invoice_date}')
    print(f'Rows Invoiced: {[row for _, row, _ in target_rows]}')
    print(f'Contracts Covered: {sorted(set(contract_no for _, _, contract_no in target_rows))}')
    print(f'\n✅ Sales file updated: {SALES_FILE}')
    print('✅ Invoice entry completed!')
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Steel Wire Invoice Entry Script',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python3 src/invoice_entry.py -c "东莞建安" -n 2026032301
  python3 src/invoice_entry.py -c "东莞建安" -n 2026032301 --invoice-date 2026.03.28
  python3 src/invoice_entry.py -c "东莞建安" -n 2026032301 --count 1
  python3 src/invoice_entry.py -c "东莞建安" -n 2026032301 --row 8
  python3 src/invoice_entry.py -c "东莞建安" --date-from 2026.03.01 --date-to 2026.03.31
  python3 src/invoice_entry.py --date-from 2026.03.01 --date-to 2026.03.31 --invoice-date 2026.03.28
        '''
    )
    parser.add_argument('-c', '--customer', help='Customer worksheet name. Optional in date-range mode.')
    parser.add_argument('-n', '--contract', help='Sales contract number.')
    parser.add_argument('--date-from', help='Delivery date start, format YYYY.MM.DD.')
    parser.add_argument('--date-to', help='Delivery date end, format YYYY.MM.DD.')
    parser.add_argument('--invoice-date', help='Invoice date, format YYYY.MM.DD. Defaults to today.')
    parser.add_argument('--count', type=int, help='How many pending rows to invoice from top to bottom.')
    parser.add_argument('--row', action='append', dest='rows', type=int, help='Specific worksheet row to invoice. Can be repeated.')
    args = parser.parse_args()

    if args.count is not None and args.rows:
        print('\n❌ Error: Use either --count or --row, not both.')
        sys.exit(1)

    if args.contract:
        if not args.customer:
            print('\n❌ Error: --customer is required when using --contract.')
            sys.exit(1)
        if args.date_from or args.date_to:
            print('\n❌ Error: Use either contract mode or date-range mode, not both.')
            sys.exit(1)
        success = enter_invoice(
            customer=args.customer,
            contract_no=args.contract,
            invoice_date=args.invoice_date,
            count=args.count,
            rows=args.rows,
        )
    else:
        if args.count is not None or args.rows:
            print('\n❌ Error: --count and --row are only supported in contract mode.')
            sys.exit(1)
        if not (args.date_from and args.date_to):
            print('\n❌ Error: Provide either --contract, or both --date-from and --date-to.')
            sys.exit(1)
        success = enter_invoice_by_delivery_range(
            customer=args.customer,
            date_from_text=args.date_from,
            date_to_text=args.date_to,
            invoice_date=args.invoice_date,
        )
    sys.exit(0 if success else 1)
