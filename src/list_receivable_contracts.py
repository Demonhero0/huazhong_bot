#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import json
import sys

from openpyxl import load_workbook

from workbook_query_utils import SALES_FILE, parse_date_text, resolve_sales_sheet_name, summarize_receivable_contracts


def main():
    parser = argparse.ArgumentParser(description='List customer receivable contracts as JSON.')
    parser.add_argument('-c', '--customer', required=True, help='Customer worksheet name.')
    parser.add_argument('-n', '--contract', help='Sales contract number.')
    parser.add_argument('--date-from', help='Sales date start, format YYYY.MM.DD.')
    parser.add_argument('--date-to', help='Sales date end, format YYYY.MM.DD.')
    parser.add_argument('--settled', default='no', choices=['yes', 'no', 'all'], help='Filter by settlement status. Default: no.')
    args = parser.parse_args()

    try:
        wb_formulas = load_workbook(SALES_FILE, data_only=False)
        resolved_customer, match_mode = resolve_sales_sheet_name(wb_formulas, args.customer)
        ws_formulas = wb_formulas[resolved_customer]
        wb_values = load_workbook(SALES_FILE, data_only=True)
        ws_values = wb_values[resolved_customer]
    except KeyError as exc:
        print(json.dumps({'ok': False, 'error': str(exc), 'workbook': SALES_FILE}, ensure_ascii=False, indent=2))
        sys.exit(1)

    contracts = summarize_receivable_contracts(
        ws_formulas=ws_formulas,
        ws_values=ws_values,
        date_from=parse_date_text(args.date_from),
        date_to=parse_date_text(args.date_to),
        settled=args.settled,
        contract_no=args.contract,
    )

    print(json.dumps({
        'ok': True,
        'workbook': SALES_FILE,
        'customer': args.customer,
        'resolved_customer': resolved_customer,
        'match_mode': match_mode,
        'filters': {
            'contract': args.contract,
            'date_from': args.date_from,
            'date_to': args.date_to,
            'settled': args.settled,
        },
        'contracts': contracts,
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
