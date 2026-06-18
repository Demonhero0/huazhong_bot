#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from workbook_query_utils import SALES_FILE, is_blank, parse_date_text, resolve_sales_sheet_name


THIN = Side(style='thin', color='000000')


def to_decimal(value, default='0.00'):
    if value is None or value == '':
        return Decimal(default)
    return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)


def iter_pending_rows(
    ws,
    date_from=None,
    date_to=None,
    require_unpaid=False,
    require_paid=False,
    require_uninvoiced=True,
):
    for row in range(5, ws.max_row + 1):
        delivery_date = ws.cell(row=row, column=8).value
        received_weight = ws.cell(row=row, column=18).value
        unit_price = ws.cell(row=row, column=19).value
        invoice_time = ws.cell(row=row, column=24).value
        receipt_date = ws.cell(row=row, column=21).value
        received_amount = ws.cell(row=row, column=22).value

        if is_blank(delivery_date) or is_blank(received_weight) or is_blank(unit_price):
            continue
        if require_uninvoiced and not is_blank(invoice_time):
            continue

        has_receipt = (not is_blank(receipt_date)) or (not is_blank(received_amount) and to_decimal(received_amount) != 0)
        if require_unpaid:
            if has_receipt:
                continue
        if require_paid and not has_receipt:
            continue

        delivery_dt = parse_date_text(str(delivery_date)) if isinstance(delivery_date, str) else None
        if delivery_dt is not None:
            if date_from and delivery_dt < date_from:
                continue
            if date_to and delivery_dt > date_to:
                continue

        yield {
            'row': row,
            'sales_date': ws.cell(row=row, column=3).value,
            'delivery_date': delivery_date,
            'spec': ws.cell(row=row, column=5).value,
            'brand': ws.cell(row=row, column=4).value,
            'received_weight': to_decimal(received_weight),
            'unit_price': to_decimal(unit_price),
            'sales_amount': (to_decimal(received_weight) * to_decimal(unit_price)).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP
            ),
        }


def autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            text = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
        ws.column_dimensions[column_letter].width = max(max_length + 2, 12)


def normalize_spec_text(value):
    text = str(value or '').strip()
    return text.replace(' ', '')


def render_row_spec(item, spec_text=None, use_row_spec=False):
    if use_row_spec:
        row_spec = normalize_spec_text(item.get('spec'))
        if row_spec:
            return row_spec
    return spec_text or ''


def format_spec_for_text(spec_value):
    text = normalize_spec_text(spec_value)
    if text.endswith('厘'):
        return f'{text[:-1]}mm'
    return text


def append_text_version(ws, start_row, rows, customer_full_name, spec_text=None, use_row_spec=False):
    lines = ['开票文字版', f'购买方：{customer_full_name}']
    for item in rows:
        row_spec = render_row_spec(item, spec_text=spec_text, use_row_spec=use_row_spec)
        lines.append(
            f'线材，规格{format_spec_for_text(row_spec)}，单位吨，数量{item["received_weight"]:.2f}，单价{item["unit_price"]:.2f}'
        )

    for idx, text in enumerate(lines):
        row = start_row + idx
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name='SimSun', size=12, bold=(idx == 0))
        cell.alignment = Alignment(horizontal='left', vertical='center')


def build_workbook(
    customer_label,
    top_text,
    spec_text,
    rows,
    use_row_spec=False,
    invoice_date_text=None,
    date_label=None,
    customer_full_name=None,
):
    wb = Workbook()
    ws = wb.active
    ws.title = '开票明细'

    ws.merge_cells('A1:D2')
    top_cell = ws['A1']
    top_cell.value = top_text
    top_cell.font = Font(name='SimSun', size=14, bold=True)
    top_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.merge_cells('E1:E2')
    invoice_date_cell = ws['E1']
    invoice_date_cell.value = invoice_date_text or ''
    invoice_date_cell.font = Font(name='SimSun', size=12, bold=True)
    invoice_date_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['日期', '规格', '订单数量/吨', '单价', '金额']
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(name='SimSun', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill('solid', fgColor='F2F2F2')
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    total_weight = Decimal('0.00')
    total_amount = Decimal('0.00')
    current_row = 4
    for item in rows:
        delivery_date = date_label or str(item['delivery_date']).strip()
        values = [
            delivery_date,
            render_row_spec(item, spec_text=spec_text, use_row_spec=use_row_spec),
            float(item['received_weight']),
            float(item['unit_price']),
            float(item['sales_amount']),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = Font(name='SimSun', size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if col in (3, 4, 5):
                cell.number_format = '0.00'
        total_weight += item['received_weight']
        total_amount += item['sales_amount']
        current_row += 1

    total_row = current_row
    totals = ['合计', '', float(total_weight), '', float(total_amount)]
    for col, value in enumerate(totals, start=1):
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = Font(name='SimSun', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        if col in (3, 5):
            cell.number_format = '0.00'

    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 24

    if customer_full_name:
        append_text_version(
            ws,
            start_row=total_row + 2,
            rows=rows,
            customer_full_name=customer_full_name,
            spec_text=spec_text,
            use_row_spec=use_row_spec,
        )

    ws.sheet_view.showGridLines = False
    autosize_columns(ws)

    return wb


def main():
    parser = argparse.ArgumentParser(description='Generate invoice detail workbook for customer pending rows.')
    parser.add_argument('-c', '--customer', required=True, help='Customer worksheet name or unique short name.')
    parser.add_argument('--date-from', help='Delivery date start, format YYYY.MM.DD.')
    parser.add_argument('--date-to', help='Delivery date end, format YYYY.MM.DD.')
    parser.add_argument('--count', type=int, help='Only export the first N matched rows.')
    parser.add_argument('--spec-text', default='线材Q195-6.5', help='Spec label written into the output workbook.')
    parser.add_argument('--top-text', help='Merged top text. Defaults to a customer-specific sentence.')
    parser.add_argument(
        '--output-dir',
        default='开票表',
        help='Output directory relative to repo root. Default: 开票表',
    )
    parser.add_argument(
        '--include-paid',
        action='store_true',
        help='Include rows that already have receipt records. Default exports unpaid rows only.',
    )
    parser.add_argument(
        '--include-invoiced',
        action='store_true',
        help='Include rows that already have invoice time. Default exports uninvoiced rows only.',
    )
    parser.add_argument(
        '--paid-only',
        action='store_true',
        help='Only export rows that already have receipt records.',
    )
    parser.add_argument(
        '--use-row-spec',
        action='store_true',
        help='Write each row using its own actual spec instead of one unified spec label.',
    )
    parser.add_argument('--invoice-date', help='Invoice date text written at top-right, format YYYY.MM.DD.')
    parser.add_argument('--date-label', help='Override every detail-row date cell with one fixed label, e.g. 5月份.')
    parser.add_argument('--customer-full-name', help='Append a copy-friendly text block using this buyer full name.')
    args = parser.parse_args()

    wb = load_workbook(SALES_FILE, data_only=True)
    resolved_customer, _ = resolve_sales_sheet_name(wb, args.customer)
    ws = wb[resolved_customer]

    date_from = parse_date_text(args.date_from) if args.date_from else None
    date_to = parse_date_text(args.date_to) if args.date_to else None

    require_unpaid = False if args.paid_only else not args.include_paid

    rows = list(
        iter_pending_rows(
            ws,
            date_from=date_from,
            date_to=date_to,
            require_unpaid=require_unpaid,
            require_paid=args.paid_only,
            require_uninvoiced=not args.include_invoiced,
        )
    )
    rows.sort(key=lambda item: (str(item['delivery_date']), item['row']))

    if args.count is not None:
        rows = rows[: args.count]

    if not rows:
        raise SystemExit(f'No matching rows found for customer: {resolved_customer}')

    if args.top_text:
        top_text = args.top_text
    else:
        top_text = f'华众开{resolved_customer}线材发票明细'

    out_dir = Path(args.output_dir) / resolved_customer
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d')
    out_path = out_dir / f'华众开{resolved_customer}线材发票明细_{timestamp}.xlsx'

    export_wb = build_workbook(
        resolved_customer,
        top_text,
        args.spec_text,
        rows,
        use_row_spec=args.use_row_spec,
        invoice_date_text=f'开票日期：{args.invoice_date}' if args.invoice_date else '',
        date_label=args.date_label,
        customer_full_name=args.customer_full_name,
    )
    export_wb.save(out_path)
    print(out_path)


if __name__ == '__main__':
    main()
