#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Steel Wire Order Entry Script

Usage:
  python3 order_entry.py -s <supplier> -p <price> -n <trucks> -b <brand> -e <spec> -t <transport> [--order-date <order_date>] [-c <customer1>] [-c <customer2>] ... [-m <payment_amount>] [-d <payment_date>]

Example:
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提 --order-date 2026.03.20
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提 -m 181151.6
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提 -c "中山富华 3420 元送到" -c "六环 3620 元送到"
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提 -c "中山富华 3420 元送到" -c "六环 3620 元送到" -d 2026.03.23 -m 181151.6
"""

from openpyxl import load_workbook
from datetime import datetime
import sys
import argparse

from workbook_query_utils import ORDER_FILE


def enter_order(supplier, unit_price, num_trucks, brand, spec, transport_method, order_date=None, customers=None, payment_date=None, payment_amount=None):
    """
    Enter supplier order information to Excel file
    
    Args:
    - supplier: Sheet name (e.g., "浙江凯航")
    - unit_price: Unit price (e.g., 3320)
    - num_trucks: Number of trucks (e.g., 2)
    - brand: Brand name (e.g., "迁安")
    - spec: Specification (e.g., "6.5 厘")
    - transport_method: Transport method (e.g., "自提")
    - order_date: Order date (e.g., "2026.03.20"), optional, defaults to today
    - customers: List of customer info strings
    - payment_date: Payment date (e.g., "2026.03.23"), optional, defaults to order_date when payment_amount is provided
    - payment_amount: Payment amount (e.g., 181151.6), optional
    """
    
    if customers is None:
        customers = []
    
    # Open file
    wb = load_workbook(ORDER_FILE)
    
    # Get worksheet
    if supplier not in wb.sheetnames:
        print(f'Error: Worksheet "{supplier}" not found!')
        print(f'Available worksheets: {wb.sheetnames}')
        return False
    
    ws = wb[supplier]
    
    # Generate contract number (today + 01)
    today = datetime.now().strftime('%Y%m%d')
    contract_no = today + '01'
    if not order_date:
        order_date = datetime.now().strftime('%Y.%m.%d')
    
    # Find last row with any data (check all columns)
    last_row = 1
    for row in range(ws.max_row, 0, -1):
        has_data = False
        for col in range(1, 20):
            if ws.cell(row=row, column=col).value is not None:
                has_data = True
                break
        if has_data:
            last_row = row
            break
    
    start_row = last_row + 1
    print(f'\n=== {supplier} Order Entry ===')
    print(f'Start row: {start_row}')
    print(f'Contract No: {contract_no}')
    print(f'Order Date: {order_date}')
    
    # Enter order information (one row per truck)
    for i in range(num_trucks):
        row = start_row + i
        
        # First truck (contract row) - fill column A
        if i == 0:
            ws.cell(row=row, column=1, value=f'{unit_price}元{num_trucks}车')
            ws.cell(row=row, column=2, value=contract_no)
            ws.cell(row=row, column=3, value=order_date)
        else:
            # Subsequent trucks - copy contract no and order date
            ws.cell(row=row, column=2, value=contract_no)
            ws.cell(row=row, column=3, value=order_date)
        
        # Fill for each truck
        ws.cell(row=row, column=4, value=brand)  # Column D: Brand
        ws.cell(row=row, column=6, value=spec)  # Column F: Specification
        ws.cell(row=row, column=7, value=transport_method)  # Column G: Transport Method
        
        # Column L: 提货金额 (Formula: J 列出厂吨数 × K 列单价)
        # 由于吨数需要后续补充，先填写单价到 K 列
        ws.cell(row=row, column=11, value=unit_price)  # Column K: 单价
        # L 列公式等待吨数填写后生效：=J 列×K 列
        ws.cell(row=row, column=12, value=f'=J{row}*K{row}')  # Column L: 提货金额
        
        # Customer info (if available)
        # P 列 (16): 客户
        if i < len(customers) and customers[i]:
            ws.cell(row=row, column=16, value=customers[i])  # Column P: Customer
        
        print(f'  Truck {i+1} (Row {row}): {brand} {spec} {transport_method}', end='')
        if i < len(customers) and customers[i]:
            print(f' -> {customers[i]} (Column P)', end='')
        print()
    
    effective_payment_date = payment_date or order_date if payment_amount is not None else None

    # Enter payment information (in first contract row)
    # M 列 (13): 付款日期，N 列 (14): 付款金额
    if payment_amount is not None:
        ws.cell(row=start_row, column=13, value=effective_payment_date)  # Column M: Payment Date
        ws.cell(row=start_row, column=14, value=payment_amount)  # Column N: Payment Amount
        print(f'  Payment: {effective_payment_date} {payment_amount} yuan (Row {start_row})')
    elif payment_date:
        print('  Warning: payment date provided without payment amount, skipping payment entry')
    
    # Fill Column O (Balance) formula
    # O 列 (15): 余款
    prev_row = start_row - 1
    prev_o_value = ws.cell(row=prev_row, column=15).value  # Column O is column 15
    
    if prev_o_value:
        for i in range(num_trucks):
            row = start_row + i
            if row == start_row:
                # Formula: =Previous_O + Current_L - Current_N
                # L 列 (12): 提货金额，N 列 (14): 付款金额，O 列 (15): 余款
                formula = f'=O{prev_row}+L{row}-N{row}'
            else:
                formula = f'=O{row-1}+L{row}-N{row}'
            ws.cell(row=row, column=15, value=formula)  # Column O: Balance

        print(f'  Balance formula filled to rows {start_row}-{start_row+num_trucks-1}')
    else:
        print(f'  Warning: Row {prev_row} Column O has no formula, skipping fill')
    
    # Save file
    wb.save(ORDER_FILE)
    print(f'\n✅ Entry completed! File saved: {ORDER_FILE}')
    return True


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Steel Wire Order Entry Script',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  # Basic order entry
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提

  # With explicit order date
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提 --order-date 2026.03.20

  # With payment amount only (payment date defaults to order date)
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提 -m 181151.6
  
  # With customer info (use -c for each customer)
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提 -c "中山富华 3420 元送到" -c "六环 3620 元送到"
  
  # With payment info
  python3 order_entry.py -s "浙江凯航" -p 3320 -n 2 -b 迁安 -e "6.5 厘" -t 自提 -c "中山富华 3420 元送到" -c "六环 3620 元送到" -d 2026.03.23 -m 181151.6
        '''
    )
    
    # Required arguments
    parser.add_argument('-s', '--supplier', required=True, help='Worksheet name (e.g., "浙江凯航")')
    parser.add_argument('-p', '--price', type=float, required=True, help='Unit price (e.g., 3320)')
    parser.add_argument('-n', '--trucks', type=int, required=True, help='Number of trucks (e.g., 2)')
    parser.add_argument('-b', '--brand', required=True, help='Brand name (e.g., "迁安")')
    parser.add_argument('-e', '--spec', required=True, help='Specification (e.g., "6.5 厘")')
    parser.add_argument('-t', '--transport', required=True, help='Transport method (e.g., "自提")')
    parser.add_argument('--order-date', help='Order date (e.g., "2026.03.20"). Defaults to today if omitted.')
    
    # Optional arguments
    parser.add_argument('-c', '--customer', action='append', default=[], help='Customer info (can be used multiple times)')
    parser.add_argument('-d', '--date', help='Payment date (e.g., "2026.03.23"). Defaults to order date when payment amount is provided.')
    parser.add_argument('-m', '--amount', type=float, help='Payment amount (e.g., 181151.6)')
    
    args = parser.parse_args()
    
    # Ensure customers list length equals num_trucks
    customers = args.customer
    while len(customers) < args.trucks:
        customers.append('')
    
    # Execute entry
    enter_order(
        supplier=args.supplier,
        unit_price=args.price,
        num_trucks=args.trucks,
        brand=args.brand,
        spec=args.spec,
        transport_method=args.transport,
        order_date=args.order_date,
        customers=customers,
        payment_date=args.date,
        payment_amount=args.amount
    )
